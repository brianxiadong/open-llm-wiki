import csv
import io
import json
import logging
import os
import re
import secrets
import shutil
from datetime import datetime, timezone

from flask import (
    Flask,
    Blueprint,
    Response,
    abort,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    stream_with_context,
    url_for,
)
from flask_login import current_user, login_required, login_user, logout_user
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from sqlalchemy import or_
from werkzeug.utils import secure_filename

from config import Config
from exceptions import MineruClientError, QdrantServiceError
from llm_client import LLMClient
from mailer import Mailer
from mineru_client import MineruClient
from models import (
    ApiToken,
    AuditLog,
    ConversationSession,
    QueryFeedback,
    QueryLog,
    Repo,
    RepoMember,
    RepoShareCode,
    Task,
    User,
    db,
    login_manager,
)
from qdrant_service import QdrantService
from task_worker import TaskWorker
from utils import (
    build_tabular_markdown_and_records,
    DEFAULT_SCHEMA_MD,
    SCHEMA_TEMPLATES,
    ensure_repo_dirs,
    extract_links,
    file_md5,
    get_backlinks,
    get_repo_path,
    list_raw_sources,
    list_wiki_pages,
    render_markdown,
    slugify,
    write_jsonl,
)
from wiki_engine import WikiEngine

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {"md", "txt", "pdf", "doc", "docx", "pptx", "png", "jpg", "jpeg",
                      "csv", "xlsx", "xls"}
EXCEL_EXTENSIONS = {"xlsx", "xls"}
MINERU_EXTENSIONS = {"pdf", "doc", "docx", "pptx", "png", "jpg", "jpeg"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TYPE_LABELS = {
    "concept": "概念",
    "guide": "指南",
    "reference": "参考",
    "overview": "概览",
    "comparison": "对比",
    "log": "日志",
    "index": "索引",
    "source": "来源",
    "entity": "实体",
    "analysis": "分析",
}
REPO_ROLE_LABELS = {
    "owner": "所有者",
    "editor": "可编辑",
    "viewer": "只读",
}
REPO_ROLE_LEVELS = {"viewer": 1, "editor": 2, "owner": 3}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _file_ext(filename: str) -> str:
    return filename.rsplit(".", 1)[1].lower() if "." in filename else ""


def _facts_records_path(base: str, source_filename: str) -> str:
    stem = os.path.splitext(source_filename)[0]
    return os.path.join(base, "facts", "records", f"{stem}.jsonl")


def _store_original_source(raw_dir: str, original_name: str, file_bytes: bytes) -> None:
    originals_dir = os.path.join(raw_dir, "originals")
    os.makedirs(originals_dir, exist_ok=True)
    with open(os.path.join(originals_dir, original_name), "wb") as f:
        f.write(file_bytes)


def _save_tabular_source(
    base: str,
    raw_dir: str,
    original_name: str,
    tables: list[dict],
) -> str:
    stem = os.path.splitext(original_name)[0]
    md_name = f"{stem}.md"
    markdown, records = build_tabular_markdown_and_records(
        source_filename=original_name,
        source_markdown_filename=md_name,
        tables=tables,
    )
    os.makedirs(os.path.join(base, "facts", "records"), exist_ok=True)
    with open(os.path.join(raw_dir, md_name), "w", encoding="utf-8") as fh:
        fh.write(markdown)
    write_jsonl(_facts_records_path(base, md_name), records)
    return md_name


def _get_repo_or_404(username: str, repo_slug: str) -> tuple:
    user = User.query.filter_by(username=username).first_or_404()
    repo = Repo.query.filter_by(user_id=user.id, slug=repo_slug).first_or_404()
    return user, repo


def _is_admin() -> bool:
    return current_user.is_authenticated and current_user.username == Config.ADMIN_USERNAME


def _require_owner(repo: Repo) -> None:
    if not current_user.is_authenticated:
        abort(403)
    if current_user.id != repo.user_id and not _is_admin():
        abort(403)


def _is_owner(repo: Repo) -> bool:
    return current_user.is_authenticated and (
        current_user.id == repo.user_id or _is_admin()
    )


def _get_repo_member(repo: Repo, user: User | None = None) -> RepoMember | None:
    if user is None:
        if not current_user.is_authenticated:
            return None
        user = current_user
    if user.id == repo.user_id:
        return None
    return RepoMember.query.filter_by(repo_id=repo.id, user_id=user.id).first()


def _get_repo_role(repo: Repo, user: User | None = None) -> str | None:
    if user is None:
        if not current_user.is_authenticated:
            return None
        user = current_user
    if _is_admin() and user.id == current_user.id:
        return "owner"
    if user.id == repo.user_id:
        return "owner"
    membership = _get_repo_member(repo, user)
    return membership.role if membership else None


def _can_edit_repo(repo: Repo, user: User | None = None) -> bool:
    role = _get_repo_role(repo, user)
    return role in {"owner", "editor"}


def _can_manage_sessions(repo: Repo) -> bool:
    if not current_user.is_authenticated:
        return False
    return repo.is_public or _get_repo_role(repo) is not None


def _can_access_repo(repo: Repo) -> bool:
    return repo.is_public or _get_repo_role(repo) is not None


def _require_editor(repo: Repo) -> None:
    if not current_user.is_authenticated:
        abort(403)
    if not _can_edit_repo(repo):
        abort(403)


def _ensure_repo_access(repo: Repo) -> None:
    if _can_access_repo(repo):
        return
    if not current_user.is_authenticated:
        return redirect(url_for("auth.login", next=request.url))
    abort(403)


def _normalize_email(email: str) -> str:
    return email.strip().lower()


def _is_valid_email(email: str) -> bool:
    return bool(EMAIL_RE.match(email))


def _reset_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(Config.SECRET_KEY, salt="password-reset")


def _verification_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(Config.SECRET_KEY, salt="email-verification")


def _build_reset_token(user: User) -> str:
    return _reset_serializer().dumps({"user_id": user.id, "password_hash": user.password_hash})


def _build_verification_token(user: User) -> str:
    return _verification_serializer().dumps({"user_id": user.id, "email": user.email})


def _verify_reset_token(token: str) -> User | None:
    try:
        payload = _reset_serializer().loads(token, max_age=Config.PASSWORD_RESET_EXPIRES)
    except (BadSignature, SignatureExpired):
        return None

    user = db.session.get(User, payload.get("user_id"))
    if user is None:
        return None
    if payload.get("password_hash") != user.password_hash:
        return None
    return user


def _verify_verification_token(token: str) -> User | None:
    try:
        payload = _verification_serializer().loads(token, max_age=Config.PASSWORD_RESET_EXPIRES)
    except (BadSignature, SignatureExpired):
        return None

    user = db.session.get(User, payload.get("user_id"))
    if user is None or not user.email:
        return None
    if payload.get("email") != user.email:
        return None
    return user


def _external_url(endpoint: str, **values) -> str:
    relative = url_for(endpoint, _external=False, **values)
    if Config.APP_BASE_URL:
        return f"{Config.APP_BASE_URL}{relative}"
    return url_for(endpoint, _external=True, **values)


def _send_verification_email(user: User) -> None:
    verify_url = _external_url("auth.verify_email", token=_build_verification_token(user))
    current_app.mailer.send_email_verification(
        user.email,
        user.username,
        verify_url,
        Config.SITE_NAME,
    )


def _audit(
    action: str,
    resource_type: str | None = None,
    resource_id: str | None = None,
    detail: str | None = None,
) -> None:
    """写入审计日志，失败静默。"""
    try:
        from models import AuditLog
        uid = current_user.id if current_user and current_user.is_authenticated else None
        uname = current_user.username if uid else None
        ip = request.remote_addr if request else None
        log = AuditLog(
            user_id=uid,
            username=uname,
            action=action,
            resource_type=resource_type,
            resource_id=str(resource_id) if resource_id is not None else None,
            detail=detail,
            ip=ip,
        )
        db.session.add(log)
        db.session.commit()
    except Exception as exc:  # noqa: BLE001
        logger.warning("audit log failed: %s", exc)


def _generate_repo_share_code() -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    while True:
        code = "KB-" + "".join(secrets.choice(alphabet) for _ in range(4)) + "-" + "".join(
            secrets.choice(alphabet) for _ in range(4)
        )
        if not RepoShareCode.query.filter_by(code=code).first():
            return code


def _enrich_sources(raw_dir: str, repo: Repo) -> list[dict]:
    sources = list_raw_sources(raw_dir)
    ingested_files: set[str] = set()
    active_tasks: dict[str, dict] = {}
    for t in Task.query.filter_by(repo_id=repo.id, type="ingest").all():
        if t.input_data:
            if t.status == "done":
                ingested_files.add(t.input_data)
            elif t.status in ("queued", "running"):
                active_tasks[t.input_data] = {
                    "status": t.status,
                    "progress": t.progress or 0,
                    "progress_msg": t.progress_msg or "",
                    "task_id": t.id,
                }

    for s in sources:
        s["id"] = s["filename"]
        kb = s["size_kb"]
        s["size_display"] = f"{kb:.1f} KB" if kb < 1024 else f"{kb / 1024:.1f} MB"
        s["ingested"] = s["filename"] in ingested_files
        task_info = active_tasks.get(s["filename"], {})
        s["task_status"] = task_info.get("status", "")
        s["task_progress"] = task_info.get("progress", 0)
        s["task_progress_msg"] = task_info.get("progress_msg", "")
        s["task_id"] = task_info.get("task_id", 0)
        filepath = os.path.join(raw_dir, s["filename"])
        try:
            mtime = os.path.getmtime(filepath)
            s["created_at"] = datetime.fromtimestamp(mtime, tz=timezone.utc)
        except OSError:
            s["created_at"] = None
    return sources


def _enrich_pages(wiki_dir: str) -> list[dict]:
    pages = list_wiki_pages(wiki_dir)
    for p in pages:
        p["slug"] = p["filename"].replace(".md", "")
    return pages


CORE_WIKI_PAGE_SLUGS = {"index", "log", "overview"}


def _default_wiki_page_content(repo_name: str, page_slug: str) -> str:
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if page_slug == "index":
        return (
            f"---\ntitle: 首页\ntype: index\nupdated: {now_str}\n---\n\n"
            f"# {repo_name}\n\n暂无内容，请上传文档并摄入。\n"
        )
    if page_slug == "log":
        return (
            "---\ntitle: Ingestion Log\ntype: log\n---\n\n"
            f"# Ingestion Log\n\n- {now_str}: 知识库创建\n"
        )
    if page_slug == "overview":
        return (
            f"---\ntitle: 概览\ntype: overview\nupdated: {now_str}\n---\n\n"
            f"# {repo_name} 概览\n\n"
            "暂无概览内容。上传文档并摄入后，此页面将自动更新。\n"
        )
    raise ValueError(f"Unsupported core wiki page: {page_slug}")


def _ensure_core_wiki_page(wiki_dir: str, repo_name: str, page_slug: str) -> tuple[str, bool]:
    filepath = os.path.join(wiki_dir, f"{page_slug}.md")
    if page_slug not in CORE_WIKI_PAGE_SLUGS or os.path.isfile(filepath):
        return filepath, False
    os.makedirs(wiki_dir, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(_default_wiki_page_content(repo_name, page_slug))
    logger.info("Recreated missing core wiki page: %s", filepath)
    return filepath, True


def _sync_repo_counts(repo: Repo, username: str) -> None:
    base = get_repo_path(Config.DATA_DIR, username, repo.slug)
    repo.source_count = len(list_raw_sources(os.path.join(base, "raw")))
    repo.page_count = len(list_wiki_pages(os.path.join(base, "wiki")))
    db.session.commit()


def _delete_repo_data(app: Flask, repo: Repo, username: str) -> str:
    repo_path = get_repo_path(Config.DATA_DIR, username, repo.slug)

    if app.qdrant:
        for fn_name in (
            "delete_collection",
            "delete_chunk_collection",
            "delete_fact_collection",
        ):
            try:
                getattr(app.qdrant, fn_name)(repo.id)
            except Exception:
                logger.warning("Failed to run %s for repo %s", fn_name, repo.id)

    RepoMember.query.filter_by(repo_id=repo.id).delete(synchronize_session=False)
    RepoShareCode.query.filter_by(repo_id=repo.id).delete(synchronize_session=False)
    Task.query.filter_by(repo_id=repo.id).delete(synchronize_session=False)
    ConversationSession.query.filter_by(repo_id=repo.id).delete(
        synchronize_session=False
    )
    QueryFeedback.query.filter_by(repo_id=repo.id).delete(
        synchronize_session=False
    )
    QueryLog.query.filter_by(repo_id=repo.id).delete(synchronize_session=False)
    db.session.delete(repo)
    return repo_path


def _delete_user_account(app: Flask, user: User) -> None:
    repo_paths = [
        _delete_repo_data(app, repo, user.username)
        for repo in Repo.query.filter_by(user_id=user.id).all()
    ]

    ApiToken.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    RepoMember.query.filter_by(user_id=user.id).delete(synchronize_session=False)
    RepoMember.query.filter_by(granted_by_user_id=user.id).update(
        {"granted_by_user_id": None}, synchronize_session=False
    )
    RepoShareCode.query.filter_by(created_by_user_id=user.id).delete(
        synchronize_session=False
    )
    ConversationSession.query.filter_by(user_id=user.id).delete(
        synchronize_session=False
    )
    QueryLog.query.filter_by(user_id=user.id).update(
        {"user_id": None}, synchronize_session=False
    )
    QueryFeedback.query.filter_by(user_id=user.id).update(
        {"user_id": None}, synchronize_session=False
    )
    AuditLog.query.filter_by(user_id=user.id).update(
        {"user_id": None}, synchronize_session=False
    )

    db.session.delete(user)
    db.session.commit()

    for repo_path in repo_paths:
        if os.path.isdir(repo_path):
            shutil.rmtree(repo_path, ignore_errors=True)

    user_dir = os.path.join(Config.DATA_DIR, user.username)
    if os.path.isdir(user_dir):
        shutil.rmtree(user_dir, ignore_errors=True)


def _purge_source_wiki(app, repo, username: str, source_filename: str) -> int:
    """Delete wiki pages derived from source_filename and their Qdrant vectors.

    Returns the number of wiki pages removed.
    """
    from utils import render_markdown
    base = get_repo_path(Config.DATA_DIR, username, repo.slug)
    wiki_dir = os.path.join(base, "wiki")
    if not os.path.isdir(wiki_dir):
        return 0

    removed = 0
    stem = os.path.splitext(source_filename)[0].lower()
    for fname in os.listdir(wiki_dir):
        if not fname.endswith(".md"):
            continue
        fpath = os.path.join(wiki_dir, fname)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                content = f.read()
            fm, _ = render_markdown(content)
            page_source = fm.get("source", "")
            # Match by source frontmatter field or filename prefix
            if (page_source and (
                    page_source == source_filename
                    or os.path.splitext(page_source)[0].lower() == stem)
                ) or fname.startswith(stem + "-") or fname == stem + ".md":
                os.remove(fpath)
                try:
                    app.qdrant.delete_page(repo.id, fname)
                except Exception:
                    pass
                try:
                    app.qdrant.delete_page_chunks(repo.id, fname)
                except Exception:
                    pass
                removed += 1
                logger.info("Purged wiki page %s (source: %s)", fname, source_filename)
        except Exception:
            pass
    return removed


def _wiki_base_url(username: str, repo_slug: str) -> str:
    raw = url_for(
        "wiki.view_page",
        username=username,
        repo_slug=repo_slug,
        page_slug="__PLACEHOLDER__",
    )
    return raw.replace("__PLACEHOLDER__", "").rstrip("/")


# ---------------------------------------------------------------------------
# App factory
# ---------------------------------------------------------------------------


def create_app() -> Flask:
    app = Flask(__name__)
    app.config.from_object(Config)
    app.config["MAX_CONTENT_LENGTH"] = Config.MAX_UPLOAD_SIZE

    db.init_app(app)
    login_manager.init_app(app)
    login_manager.login_view = "auth.login"
    login_manager.login_message = "请先登录"

    app.llm = LLMClient(
        Config.LLM_API_BASE,
        Config.LLM_API_KEY,
        Config.LLM_MODEL,
        Config.LLM_MAX_TOKENS,
    )
    app.mailer = Mailer(
        Config.MAIL_HOST,
        Config.MAIL_PORT,
        Config.MAIL_USERNAME,
        Config.MAIL_PASSWORD,
        use_ssl=Config.MAIL_USE_SSL,
        default_from=Config.MAIL_FROM,
    )
    app.mineru = MineruClient(Config.MINERU_API_URL, Config.MINERU_TIMEOUT)
    try:
        app.qdrant = QdrantService(
            Config.QDRANT_URL,
            Config.EMBEDDING_API_BASE,
            Config.EMBEDDING_API_KEY,
            Config.EMBEDDING_MODEL,
            Config.EMBEDDING_DIMENSIONS,
        )
    except Exception:
        app.qdrant = None
        logging.warning("Qdrant service unavailable, running in degraded mode")
    app.wiki_engine = WikiEngine(app.llm, app.qdrant, Config.DATA_DIR)

    # Query trace logger: daily JSONL files under DATA_DIR/logs/
    from utils import QueryTraceLogger
    app.query_trace_logger = QueryTraceLogger(os.path.join(Config.DATA_DIR, "logs"))

    @app.context_processor
    def inject_admin():
        return {"admin_username": Config.ADMIN_USERNAME, "site_name": Config.SITE_NAME}

    @app.before_request
    def _check_api_token():
        """Bearer token 认证，供脚本和 API 集成调用。"""
        if current_user.is_authenticated:
            return
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return
        raw = auth[7:].strip()
        if not raw:
            return
        import hashlib
        from models import ApiToken
        h = hashlib.sha256(raw.encode()).hexdigest()
        token = ApiToken.query.filter_by(token_hash=h, is_active=True).first()
        if token:
            login_user(token.user, remember=False)
            token.last_used_at = datetime.now(timezone.utc)
            db.session.commit()

    _register_routes(app)
    _register_error_handlers(app)

    os.makedirs(Config.DATA_DIR, exist_ok=True)

    if not app.config.get("TESTING"):
        worker = TaskWorker(app)
        worker.start()
        app.task_worker = worker

    return app


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------


def _register_error_handlers(app: Flask) -> None:
    @app.errorhandler(404)
    def not_found(_e):
        return render_template("errors/404.html"), 404

    @app.errorhandler(500)
    def server_error(_e):
        return render_template("errors/500.html"), 500


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


def _register_routes(app: Flask) -> None:

    # ── Root ──────────────────────────────────────────────────────────

    @app.route("/")
    def index():
        if current_user.is_authenticated:
            return redirect(
                url_for("repo.list_repos", username=current_user.username)
            )
        return redirect(url_for("auth.login"))

    @app.route("/guide")
    def guide():
        return render_template("guide.html")

    @app.route("/health")
    def health():
        checks: dict[str, str] = {}

        try:
            db.session.execute(db.text("SELECT 1"))
            checks["mysql"] = "ok"
        except Exception as exc:
            checks["mysql"] = f"error: {exc}"

        if app.qdrant:
            try:
                app.qdrant._qdrant.get_collections()
                checks["qdrant"] = "ok"
            except Exception as exc:
                checks["qdrant"] = f"error: {exc}"
        else:
            checks["qdrant"] = "unavailable"

        try:
            checks["mineru"] = "ok" if app.mineru.health_check() else "error"
        except Exception as exc:
            checks["mineru"] = f"error: {exc}"

        if app.qdrant:
            try:
                app.qdrant._embed("health-check")
                checks["embedding"] = "ok"
            except Exception as exc:
                checks["embedding"] = f"error: {exc}"
        else:
            checks["embedding"] = "unavailable"

        all_ok = all(v == "ok" for v in checks.values())
        return (
            jsonify(status="ok" if all_ok else "degraded", checks=checks),
            200 if all_ok else 503,
        )

    # ── Auth ──────────────────────────────────────────────────────────

    auth_bp = Blueprint("auth", __name__)

    @auth_bp.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(
                url_for("repo.list_repos", username=current_user.username)
            )
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            email = _normalize_email(username)
            user = User.query.filter(
                or_(User.username == username, User.email == email)
            ).first()
            if user and user.check_password(password):
                if user.email and not user.email_verified:
                    if current_app.mailer.enabled:
                        try:
                            _send_verification_email(user)
                            flash("账号尚未完成邮箱验证，已重新发送验证邮件，请查收邮箱", "warning")
                        except Exception as exc:
                            logger.exception("send verification mail failed")
                            flash(f"账号尚未完成邮箱验证，且重发验证邮件失败: {exc}", "error")
                    else:
                        flash("账号尚未完成邮箱验证，且邮件服务未配置，请联系管理员", "error")
                    return render_template("auth/login.html")
                login_user(user)
                _audit("login")
                next_url = request.args.get("next")
                return redirect(
                    next_url
                    or url_for("repo.list_repos", username=user.username)
                )
            flash("用户名或密码错误", "error")
        return render_template("auth/login.html")

    @auth_bp.route("/register", methods=["GET", "POST"])
    def register():
        if current_user.is_authenticated:
            return redirect(
                url_for("repo.list_repos", username=current_user.username)
            )
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            email = _normalize_email(request.form.get("email", ""))
            display_name = request.form.get("display_name", "").strip()
            password = request.form.get("password", "")
            confirm = request.form.get("confirm_password", "")

            if not username or not email or not password:
                flash("用户名、邮箱和密码不能为空", "error")
                return render_template("auth/register.html")
            if not re.match(r"^[a-zA-Z0-9_]+$", username):
                flash("用户名只能包含字母、数字和下划线", "error")
                return render_template("auth/register.html")
            if not _is_valid_email(email):
                flash("请输入有效的邮箱地址", "error")
                return render_template("auth/register.html")
            if len(password) < 8:
                flash("密码长度不能少于 8 位", "error")
                return render_template("auth/register.html")
            if password != confirm:
                flash("两次输入的密码不一致", "error")
                return render_template("auth/register.html")
            if User.query.filter_by(username=username).first():
                flash("用户名已存在", "error")
                return render_template("auth/register.html")
            if User.query.filter_by(email=email).first():
                flash("邮箱已被使用", "error")
                return render_template("auth/register.html")
            if not current_app.mailer.enabled:
                flash("邮件服务未配置，暂时无法完成注册", "error")
                return render_template("auth/register.html")

            user = User(username=username, email=email, display_name=display_name or None)
            user.set_password(password)
            db.session.add(user)
            try:
                db.session.flush()
                _send_verification_email(user)
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                logger.exception("send verification mail failed on register")
                flash(f"验证邮件发送失败: {exc}", "error")
                return render_template("auth/register.html")

            flash("注册成功，请先查收验证邮件并完成邮箱认证后再登录", "success")
            return redirect(url_for("auth.login"))
        return render_template("auth/register.html")

    @auth_bp.route("/verify-email/<token>", methods=["GET"])
    def verify_email(token):
        user = _verify_verification_token(token)
        if user is None:
            flash("验证链接无效或已过期，请重新登录后获取新的验证邮件", "error")
            return redirect(url_for("auth.login"))
        if user.email_verified:
            flash("邮箱已完成验证，请直接登录", "success")
            return redirect(url_for("auth.login"))

        user.email_verified = True
        user.email_verified_at = datetime.now(timezone.utc)
        db.session.commit()
        _audit("verify_email", "user", user.id, user.email)
        flash("邮箱验证成功，请登录", "success")
        return redirect(url_for("auth.login"))

    @auth_bp.route("/forgot-password", methods=["GET", "POST"])
    def forgot_password():
        if current_user.is_authenticated:
            return redirect(
                url_for("repo.list_repos", username=current_user.username)
            )
        if request.method == "POST":
            email = _normalize_email(request.form.get("email", ""))
            user = User.query.filter_by(email=email).first() if email else None
            if not current_app.mailer.enabled:
                flash("邮件服务未配置，暂时无法发送找回密码邮件", "error")
                return render_template("auth/forgot_password.html")
            if user and user.email:
                try:
                    reset_url = _external_url("auth.reset_password", token=_build_reset_token(user))
                    current_app.mailer.send_password_reset(
                        user.email,
                        user.username,
                        reset_url,
                        Config.SITE_NAME,
                    )
                except Exception as exc:
                    logger.exception("send reset mail failed")
                    flash(f"邮件发送失败: {exc}", "error")
                    return render_template("auth/forgot_password.html")
            flash("如果邮箱存在对应账号，重置链接已经发送", "success")
            return redirect(url_for("auth.login"))
        return render_template("auth/forgot_password.html")

    @auth_bp.route("/reset-password/<token>", methods=["GET", "POST"])
    def reset_password(token):
        if current_user.is_authenticated:
            return redirect(
                url_for("repo.list_repos", username=current_user.username)
            )
        user = _verify_reset_token(token)
        if user is None:
            flash("重置链接无效或已过期", "error")
            return redirect(url_for("auth.forgot_password"))
        if request.method == "POST":
            password = request.form.get("password", "")
            confirm = request.form.get("confirm_password", "")
            if len(password) < 8:
                flash("密码长度不能少于 8 位", "error")
                return render_template("auth/reset_password.html", token=token)
            if password != confirm:
                flash("两次输入的密码不一致", "error")
                return render_template("auth/reset_password.html", token=token)
            user.set_password(password)
            db.session.commit()
            flash("密码已重置，请重新登录", "success")
            return redirect(url_for("auth.login"))
        return render_template("auth/reset_password.html", token=token)

    @auth_bp.route("/logout")
    def logout():
        _audit("logout")
        logout_user()
        return redirect(url_for("auth.login"))

    app.register_blueprint(auth_bp)

    # ── User ──────────────────────────────────────────────────────────

    user_bp = Blueprint("user", __name__, url_prefix="/user")

    @user_bp.route("/settings", methods=["GET", "POST"])
    @login_required
    def settings():
        if request.method == "POST":
            action = request.form.get("action")
            if action == "update_profile":
                email = _normalize_email(request.form.get("email", ""))
                if not email:
                    flash("邮箱不能为空", "error")
                    return redirect(url_for("user.settings"))
                if not _is_valid_email(email):
                    flash("请输入有效的邮箱地址", "error")
                    return redirect(url_for("user.settings"))
                existed = User.query.filter(
                    User.email == email,
                    User.id != current_user.id,
                ).first()
                if existed:
                    flash("邮箱已被其他账号使用", "error")
                    return redirect(url_for("user.settings"))
                current_user.email = email
                current_user.display_name = (
                    request.form.get("display_name", "").strip() or None
                )
                db.session.commit()
                flash("个人信息已更新", "success")
            elif action == "change_password":
                old_pw = request.form.get("old_password", "")
                new_pw = request.form.get("new_password", "")
                confirm = request.form.get("confirm_password", "")
                if not current_user.check_password(old_pw):
                    flash("当前密码错误", "error")
                elif new_pw != confirm:
                    flash("两次输入的新密码不一致", "error")
                elif len(new_pw) < 8:
                    flash("新密码长度不能少于 8 位", "error")
                else:
                    current_user.set_password(new_pw)
                    db.session.commit()
                    flash("密码修改成功", "success")
            elif action == "delete_account":
                confirm_username = request.form.get("confirm_username", "").strip()
                password = request.form.get("delete_password", "")
                if confirm_username != current_user.username:
                    flash("请输入当前用户名以确认删除账号", "error")
                elif not current_user.check_password(password):
                    flash("当前密码错误，无法删除账号", "error")
                else:
                    app_obj = current_app._get_current_object()
                    username = current_user.username
                    user_id = current_user.id
                    _audit("delete_account", "user", user_id, username)
                    _delete_user_account(app_obj, current_user)
                    logout_user()
                    flash("账号及其关联知识库已删除", "success")
                    return redirect(url_for("auth.login"))
            return redirect(url_for("user.settings"))
        return render_template("user/settings.html")

    @user_bp.route("/settings/tokens", methods=["GET"])
    @login_required
    def list_tokens():
        from models import ApiToken
        tokens = (
            ApiToken.query.filter_by(user_id=current_user.id)
            .order_by(ApiToken.created_at.desc())
            .all()
        )
        return render_template("user/tokens.html", tokens=tokens)

    @user_bp.route("/settings/tokens/create", methods=["POST"])
    @login_required
    def create_token():
        import hashlib
        import secrets
        from models import ApiToken
        name = request.form.get("name", "").strip()
        if not name:
            flash("Token 名称不能为空", "error")
            return redirect(url_for("user.list_tokens"))
        raw = secrets.token_urlsafe(32)
        h = hashlib.sha256(raw.encode()).hexdigest()
        t = ApiToken(user_id=current_user.id, name=name, token_hash=h)
        db.session.add(t)
        db.session.commit()
        _audit("create_api_token", "api_token", t.id, name)
        flash(f"Token 已创建（只显示一次，请立即复制）：{raw}", "success")
        return redirect(url_for("user.list_tokens"))

    @user_bp.route("/settings/tokens/<int:token_id>/revoke", methods=["POST"])
    @login_required
    def revoke_token(token_id):
        from models import ApiToken
        t = ApiToken.query.filter_by(id=token_id, user_id=current_user.id).first_or_404()
        t.is_active = False
        db.session.commit()
        _audit("revoke_api_token", "api_token", token_id)
        flash("Token 已吊销", "success")
        return redirect(url_for("user.list_tokens"))

    app.register_blueprint(user_bp)

    # ── Repo ──────────────────────────────────────────────────────────

    repo_bp = Blueprint("repo", __name__)

    @repo_bp.route("/<username>")
    def list_repos(username):
        user = User.query.filter_by(username=username).first_or_404()
        page_is_self = (
            current_user.is_authenticated
            and (current_user.id == user.id or _is_admin())
        )
        all_owned_repos = (
            Repo.query.filter_by(user_id=user.id)
            .order_by(Repo.updated_at.desc())
            .all()
        )
        repos = (
            all_owned_repos
            if page_is_self
            else [repo for repo in all_owned_repos if _can_access_repo(repo)]
        )
        shared_repos = []
        if page_is_self:
            memberships = (
                RepoMember.query.filter_by(user_id=current_user.id)
                .order_by(RepoMember.created_at.desc())
                .all()
            )
            for membership in memberships:
                repo = db.session.get(Repo, membership.repo_id)
                if repo is None or repo.user_id == current_user.id:
                    continue
                repo.access_role = membership.role
                repo.access_role_label = REPO_ROLE_LABELS.get(membership.role, membership.role)
                repo.owner_username = repo.user.username
                repo.is_shared_repo = True
                shared_repos.append(repo)
        # Attach active task count to each repo for the card UI
        for repo in repos:
            role = _get_repo_role(repo)
            repo.access_role = role or ("viewer" if repo.is_public else None)
            repo.access_role_label = (
                REPO_ROLE_LABELS.get(repo.access_role, "公开访客")
                if repo.access_role
                else "公开访客"
            )
            repo.owner_username = user.username
            repo.is_shared_repo = False
            repo.active_task_count = Task.query.filter(
                Task.repo_id == repo.id,
                Task.status.in_(["queued", "running"]),
            ).count()
        for repo in shared_repos:
            repo.active_task_count = Task.query.filter(
                Task.repo_id == repo.id,
                Task.status.in_(["queued", "running"]),
            ).count()
        return render_template(
            "repo/list.html",
            username=username,
            repos=repos,
            shared_repos=shared_repos,
            page_is_self=page_is_self,
        )

    @repo_bp.route("/repos/join", methods=["POST"])
    @login_required
    def join_by_code():
        raw_code = request.form.get("access_code", "").strip().upper()
        if not raw_code:
            flash("请输入访问码", "error")
            return redirect(url_for("repo.list_repos", username=current_user.username))

        share_code = RepoShareCode.query.filter_by(code=raw_code).first()
        if share_code is None or not share_code.is_active:
            flash("访问码无效或已停用", "error")
            return redirect(url_for("repo.list_repos", username=current_user.username))

        repo = db.session.get(Repo, share_code.repo_id)
        if repo is None:
            flash("访问码关联的知识库不存在", "error")
            return redirect(url_for("repo.list_repos", username=current_user.username))
        if repo.user_id == current_user.id:
            flash("这是你自己的知识库，无需加入", "info")
            return redirect(url_for("repo.list_repos", username=current_user.username))

        existing = RepoMember.query.filter_by(
            repo_id=repo.id, user_id=current_user.id
        ).first()
        if existing:
            flash("该共享知识库已在你的列表中", "info")
            return redirect(url_for("repo.list_repos", username=current_user.username))

        membership = RepoMember(
            repo_id=repo.id,
            user_id=current_user.id,
            role=share_code.role,
            granted_by_user_id=share_code.created_by_user_id,
            share_code_id=share_code.id,
        )
        db.session.add(membership)
        share_code.use_count += 1
        db.session.commit()
        flash(
            f"已加入共享知识库「{repo.name}」({REPO_ROLE_LABELS.get(share_code.role, share_code.role)})",
            "success",
        )
        return redirect(url_for("repo.list_repos", username=current_user.username))

    @repo_bp.route("/shared-repos/<int:repo_id>/leave", methods=["POST"])
    @login_required
    def leave_shared_repo(repo_id):
        membership = RepoMember.query.filter_by(
            repo_id=repo_id, user_id=current_user.id
        ).first_or_404()
        repo = db.session.get(Repo, membership.repo_id)
        db.session.delete(membership)
        db.session.commit()
        flash(
            f"已退出共享知识库「{repo.name if repo else '未知知识库'}」",
            "success",
        )
        return redirect(url_for("repo.list_repos", username=current_user.username))

    @repo_bp.route("/repos/new", methods=["GET", "POST"])
    @login_required
    def new_repo():
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            slug_raw = request.form.get("slug", "").strip()
            description = request.form.get("description", "").strip()

            if not name:
                flash("知识库名称不能为空", "error")
                return render_template("repo/new.html", schema_templates=SCHEMA_TEMPLATES)

            slug = slug_raw or slugify(name)
            if not slug:
                flash("无法生成有效的 URL 标识", "error")
                return render_template("repo/new.html", schema_templates=SCHEMA_TEMPLATES)

            if Repo.query.filter_by(user_id=current_user.id, slug=slug).first():
                flash("该标识已被使用", "error")
                return render_template("repo/new.html", schema_templates=SCHEMA_TEMPLATES)

            repo = Repo(
                user_id=current_user.id,
                name=name,
                slug=slug,
                description=description,
            )
            db.session.add(repo)
            db.session.commit()

            base = ensure_repo_dirs(Config.DATA_DIR, current_user.username, slug)
            wiki_dir = os.path.join(base, "wiki")

            schema_template_key = request.form.get("schema_template", "default")
            _, schema_content = SCHEMA_TEMPLATES.get(schema_template_key, SCHEMA_TEMPLATES["default"])
            with open(os.path.join(wiki_dir, "schema.md"), "w", encoding="utf-8") as f:
                f.write(schema_content)

            for page_slug in sorted(CORE_WIKI_PAGE_SLUGS):
                _ensure_core_wiki_page(wiki_dir, name, page_slug)

            if app.qdrant:
                try:
                    app.qdrant.ensure_collection(repo.id)
                except Exception:
                    logger.warning(
                        "Failed to create Qdrant collection for repo %s", repo.id
                    )

            _sync_repo_counts(repo, current_user.username)
            flash("知识库创建成功", "success")
            return redirect(
                url_for(
                    "repo.dashboard",
                    username=current_user.username,
                    repo_slug=slug,
                )
            )
        return render_template("repo/new.html", schema_templates=SCHEMA_TEMPLATES)

    @repo_bp.route("/<username>/<repo_slug>")
    def dashboard(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return redirect(url_for("auth.login", next=request.url))
            abort(403)
        base = get_repo_path(Config.DATA_DIR, username, repo_slug)
        wiki_dir = os.path.join(base, "wiki")
        _, overview_created = _ensure_core_wiki_page(wiki_dir, repo.name, "overview")
        pages = _enrich_pages(wiki_dir)
        raw_dir = os.path.join(base, "raw")
        sources = _enrich_sources(raw_dir, repo) if os.path.isdir(raw_dir) else []
        if overview_created or repo.page_count != len(pages) or repo.source_count != len(sources):
            _sync_repo_counts(repo, username)

        page_content = None
        active_page = None
        overview_path = os.path.join(wiki_dir, "overview.md")
        if os.path.isfile(overview_path):
            with open(overview_path, "r", encoding="utf-8") as f:
                raw = f.read()
            _, page_content = render_markdown(raw, _wiki_base_url(username, repo_slug))
            for p in pages:
                if p["slug"] == "overview":
                    active_page = p
                    break

        # README
        readme_html = ""
        readme_path = os.path.join(base, "README.md")
        if os.path.isfile(readme_path):
            try:
                with open(readme_path, "r", encoding="utf-8") as f:
                    raw_readme = f.read()
                _, readme_html = render_markdown(raw_readme, _wiki_base_url(username, repo_slug))
            except Exception:
                pass

        return render_template(
            "repo/dashboard.html",
            username=username,
            repo=repo,
            pages=pages,
            sources=sources,
            page_content=page_content,
            active_page=active_page,
            is_owner=_is_owner(repo),
            can_edit=_can_edit_repo(repo),
            can_manage_sessions=_can_manage_sessions(repo),
            readme_html=readme_html,
        )

    @repo_bp.route(
        "/<username>/<repo_slug>/settings",
        methods=["GET", "POST"],
        endpoint="settings",
    )
    @login_required
    def repo_settings(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)

        base = get_repo_path(Config.DATA_DIR, username, repo_slug)
        schema_path = os.path.join(base, "wiki", "schema.md")

        if request.method == "POST":
            action = request.form.get("action")
            if action == "update_info":
                repo.name = request.form.get("name", "").strip() or repo.name
                repo.description = request.form.get("description", "").strip()
                repo.is_public = "is_public" in request.form
                db.session.commit()
                flash("设置已保存", "success")
            elif action == "update_schema":
                content = request.form.get("schema_content", "")
                os.makedirs(os.path.dirname(schema_path), exist_ok=True)
                with open(schema_path, "w", encoding="utf-8") as f:
                    f.write(content)
                flash("Schema 已保存", "success")
            elif action == "update_readme":
                readme_content = request.form.get("readme", "")
                readme_path = os.path.join(base, "README.md")
                with open(readme_path, "w", encoding="utf-8") as f:
                    f.write(readme_content)
                flash("README 已保存", "success")
            elif action == "create_share_code":
                role = request.form.get("share_role", "viewer").strip()
                if role not in {"viewer", "editor"}:
                    flash("共享权限无效", "error")
                else:
                    share_code = RepoShareCode(
                        repo_id=repo.id,
                        code=_generate_repo_share_code(),
                        role=role,
                        created_by_user_id=current_user.id,
                    )
                    db.session.add(share_code)
                    db.session.commit()
                    flash(
                        f"已生成访问码：{share_code.code}（{REPO_ROLE_LABELS.get(role, role)}）",
                        "success",
                    )
            return redirect(
                url_for("repo.settings", username=username, repo_slug=repo_slug)
            )

        schema_content = ""
        if os.path.isfile(schema_path):
            with open(schema_path, "r", encoding="utf-8") as f:
                schema_content = f.read()

        readme_content = ""
        readme_path = os.path.join(base, "README.md")
        if os.path.isfile(readme_path):
            with open(readme_path, "r", encoding="utf-8") as f:
                readme_content = f.read()

        repo_members = (
            RepoMember.query.filter_by(repo_id=repo.id)
            .order_by(RepoMember.created_at.asc())
            .all()
        )
        repo_share_codes = (
            RepoShareCode.query.filter_by(repo_id=repo.id)
            .order_by(RepoShareCode.created_at.desc())
            .all()
        )

        return render_template(
            "repo/settings.html",
            username=username,
            repo=repo,
            schema_content=schema_content,
            readme_content=readme_content,
            repo_members=repo_members,
            repo_share_codes=repo_share_codes,
            role_labels=REPO_ROLE_LABELS,
        )

    @repo_bp.route("/<username>/<repo_slug>/members/<int:member_id>/delete", methods=["POST"])
    @login_required
    def remove_member(username, repo_slug, member_id):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)
        membership = RepoMember.query.filter_by(
            id=member_id, repo_id=repo.id
        ).first_or_404()
        member_user = db.session.get(User, membership.user_id)
        db.session.delete(membership)
        db.session.commit()
        flash(
            f"已移除成员「{member_user.username if member_user else member_id}」",
            "success",
        )
        return redirect(url_for("repo.settings", username=username, repo_slug=repo_slug))

    @repo_bp.route("/<username>/<repo_slug>/share-codes/<int:code_id>/disable", methods=["POST"])
    @login_required
    def disable_share_code(username, repo_slug, code_id):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)
        share_code = RepoShareCode.query.filter_by(
            id=code_id, repo_id=repo.id
        ).first_or_404()
        share_code.is_active = False
        db.session.commit()
        flash(f"访问码 {share_code.code} 已停用", "success")
        return redirect(url_for("repo.settings", username=username, repo_slug=repo_slug))

    @repo_bp.route("/<username>/<repo_slug>/delete", methods=["POST"])
    @login_required
    def delete(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)

        repo_path = _delete_repo_data(app, repo, username)
        db.session.commit()

        if os.path.isdir(repo_path):
            shutil.rmtree(repo_path, ignore_errors=True)

        _audit("delete_repo", "repo", repo.id, repo_slug)
        flash("知识库已删除", "success")
        return redirect(url_for("repo.list_repos", username=username))

    @repo_bp.route("/<username>/search")
    def global_search(username):
        user = User.query.filter_by(username=username).first_or_404()
        if not current_user.is_authenticated or current_user.id != user.id:
            abort(403)
        query_text = request.args.get("q", "").strip()
        results = []
        if query_text:
            query_lower = query_text.lower()
            repos = Repo.query.filter_by(user_id=user.id).all()
            for repo in repos:
                wiki_dir = os.path.join(
                    get_repo_path(Config.DATA_DIR, username, repo.slug), "wiki"
                )
                if not os.path.isdir(wiki_dir):
                    continue
                repo_results = []
                for filename in sorted(os.listdir(wiki_dir)):
                    if not filename.endswith(".md"):
                        continue
                    filepath = os.path.join(wiki_dir, filename)
                    try:
                        with open(filepath, "r", encoding="utf-8") as f:
                            content = f.read()
                    except OSError:
                        continue
                    content_lower = content.lower()
                    if query_lower not in content_lower:
                        continue
                    fm, _ = render_markdown(content)
                    idx = content_lower.find(query_lower)
                    start = max(0, idx - 80)
                    end = min(len(content), idx + len(query_text) + 80)
                    snippet = content[start:end].replace("\n", " ")
                    if start > 0:
                        snippet = "…" + snippet
                    if end < len(content):
                        snippet += "…"
                    repo_results.append({
                        "slug": filename.replace(".md", ""),
                        "title": fm.get("title", filename.replace(".md", "")),
                        "type": fm.get("type", "unknown"),
                        "snippet": snippet,
                        "match_count": content_lower.count(query_lower),
                    })
                if repo_results:
                    repo_results.sort(key=lambda r: r["match_count"], reverse=True)
                    results.append({
                        "repo": repo,
                        "pages": repo_results,
                        "total_matches": sum(r["match_count"] for r in repo_results),
                    })
            results.sort(key=lambda r: r["total_matches"], reverse=True)

        return render_template(
            "user/search.html",
            username=username,
            query=query_text,
            results=results,
            profile_user=user,
        )

    @repo_bp.route("/<username>/<repo_slug>/import.zip", methods=["POST"])
    @login_required
    def import_zip(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)
        f = request.files.get("file")
        if not f or not f.filename.endswith(".zip"):
            flash("请上传 .zip 文件", "error")
            return redirect(url_for("repo.settings", username=username, repo_slug=repo_slug))

        import zipfile
        import uuid as _uuid
        mode = request.form.get("mode", "merge")
        base = get_repo_path(Config.DATA_DIR, username, repo_slug)
        tmp = os.path.join(base, f"temp_import_{_uuid.uuid4().hex[:8]}")
        os.makedirs(tmp, exist_ok=True)
        try:
            zip_path = os.path.join(tmp, "upload.zip")
            f.save(zip_path)
            with zipfile.ZipFile(zip_path, "r") as zf:
                zf.extractall(tmp)
            os.remove(zip_path)

            wiki_src = os.path.join(tmp, "wiki")
            raw_src = os.path.join(tmp, "raw")
            if not os.path.isdir(wiki_src):
                for sub in os.listdir(tmp):
                    candidate = os.path.join(tmp, sub, "wiki")
                    if os.path.isdir(candidate):
                        wiki_src = candidate
                        raw_src = os.path.join(tmp, sub, "raw")
                        break

            if not os.path.isdir(wiki_src):
                flash("ZIP 中未找到 wiki/ 目录", "error")
                return redirect(url_for("repo.settings", username=username, repo_slug=repo_slug))

            wiki_dst = os.path.join(base, "wiki")
            raw_dst = os.path.join(base, "raw")
            os.makedirs(wiki_dst, exist_ok=True)
            os.makedirs(raw_dst, exist_ok=True)

            if mode == "replace":
                for fn in os.listdir(wiki_dst):
                    fp = os.path.join(wiki_dst, fn)
                    if os.path.isfile(fp):
                        os.remove(fp)

            imported_wiki = 0
            for fn in os.listdir(wiki_src):
                if fn.endswith(".md"):
                    shutil.copy2(os.path.join(wiki_src, fn), os.path.join(wiki_dst, fn))
                    imported_wiki += 1
            if os.path.isdir(raw_src):
                for fn in os.listdir(raw_src):
                    src_fp = os.path.join(raw_src, fn)
                    if os.path.isfile(src_fp):
                        shutil.copy2(src_fp, os.path.join(raw_dst, fn))

            task = Task(repo_id=repo.id, type="rebuild_index", status="queued", input_data="import_zip")
            db.session.add(task)
            repo.page_count = len(list_wiki_pages(wiki_dst))
            db.session.commit()
            _audit("import_zip", "repo", repo.id, f"mode={mode} pages={imported_wiki}")
            flash(f"导入完成，共 {imported_wiki} 个 Wiki 页面。索引重建已加入队列。", "success")
        except Exception as exc:
            logger.exception("import_zip failed")
            flash(f"导入失败：{exc}", "error")
        finally:
            shutil.rmtree(tmp, ignore_errors=True)
        return redirect(url_for("repo.dashboard", username=username, repo_slug=repo_slug))

    app.register_blueprint(repo_bp)

    # ── Wiki ──────────────────────────────────────────────────────────

    wiki_bp = Blueprint("wiki", __name__)

    @wiki_bp.route("/<username>/<repo_slug>/wiki/<page_slug>")
    def view_page(username, repo_slug, page_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return redirect(url_for("auth.login", next=request.url))
            abort(403)
        wiki_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "wiki"
        )
        filepath, page_created = _ensure_core_wiki_page(wiki_dir, repo.name, page_slug)
        if page_created:
            _sync_repo_counts(repo, username)
        if not os.path.isfile(filepath):
            abort(404)

        with open(filepath, "r", encoding="utf-8") as f:
            raw = f.read()

        fm, html = render_markdown(raw, _wiki_base_url(username, repo_slug))

        try:
            mtime = datetime.fromtimestamp(
                os.path.getmtime(filepath), tz=timezone.utc
            )
        except OSError:
            mtime = None

        page_sources = []
        if fm.get("source"):
            src_name = fm["source"]
            page_sources.append({"id": src_name, "filename": src_name})

        page = {
            "title": fm.get("title", page_slug),
            "type": fm.get("type", "unknown"),
            "type_label": TYPE_LABELS.get(fm.get("type", ""), fm.get("type", "")),
            "created_at": mtime,
            "updated_at": mtime,
            "sources": page_sources,
        }

        backlinks = get_backlinks(wiki_dir, page_slug)
        for bl in backlinks:
            bl["slug"] = bl["filename"].replace(".md", "")

        return render_template(
            "wiki/page.html",
            username=username,
            repo=repo,
            page=page,
            page_slug=page_slug,
            content=html,
            backlinks=backlinks,
            is_owner=_is_owner(repo),
            can_edit=_can_edit_repo(repo),
        )

    @wiki_bp.route("/<username>/<repo_slug>/graph")
    def graph(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return redirect(url_for("auth.login", next=request.url))
            abort(403)
        wiki_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "wiki"
        )
        pages = _enrich_pages(wiki_dir)
        page_slugs = {p["slug"] for p in pages}

        nodes = []
        links = []
        for p in pages:
            nodes.append(
                {
                    "id": p["slug"],
                    "title": p["title"],
                    "type": p.get("type", "unknown"),
                }
            )
            filepath = os.path.join(wiki_dir, p["filename"])
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    content = f.read()
            except OSError:
                continue
            for target in extract_links(content):
                if target in page_slugs:
                    links.append({"source": p["slug"], "target": target})

        return render_template(
            "wiki/graph.html",
            username=username,
            repo=repo,
            graph_data={"nodes": nodes, "links": links},
        )

    @wiki_bp.route("/<username>/<repo_slug>/log")
    def log_page(username, repo_slug):
        _get_repo_or_404(username, repo_slug)
        return redirect(
            url_for(
                "wiki.view_page",
                username=username,
                repo_slug=repo_slug,
                page_slug="log",
            )
        )

    @wiki_bp.route("/<username>/<repo_slug>/wiki/pages")
    def wiki_pages(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            abort(403)
        wiki_dir = os.path.join(get_repo_path(Config.DATA_DIR, username, repo_slug), "wiki")
        all_pages = list_wiki_pages(wiki_dir) if os.path.isdir(wiki_dir) else []
        # add slug derived from filename
        for p in all_pages:
            p["slug"] = p["filename"].replace(".md", "")
        # group by type
        groups: dict[str, list] = {}
        for p in all_pages:
            t = p.get("type", "other")
            groups.setdefault(t, []).append(p)
        type_order = ["overview", "concept", "reference", "comparison", "guide", "changelog", "other"]
        type_labels = {
            "overview": "概览", "concept": "概念", "reference": "参考",
            "comparison": "对比", "guide": "指南", "changelog": "变更日志", "other": "其他",
        }
        sorted_groups = [(t, type_labels.get(t, t), groups[t]) for t in type_order if t in groups]
        for t, pages in groups.items():
            if t not in type_order:
                sorted_groups.append((t, t, pages))
        return render_template(
            "wiki/pages.html",
            username=username, repo=repo,
            all_pages=all_pages,
            sorted_groups=sorted_groups,
            is_owner=_is_owner(repo),
            can_edit=_can_edit_repo(repo),
        )

    @wiki_bp.route("/<username>/<repo_slug>/wiki/<page_slug>/edit", methods=["GET", "POST"])
    @login_required
    def edit_page(username, repo_slug, page_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        wiki_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "wiki"
        )
        filepath, page_created = _ensure_core_wiki_page(wiki_dir, repo.name, page_slug)
        if page_created:
            _sync_repo_counts(repo, username)
        if not os.path.isfile(filepath):
            abort(404)

        if request.method == "POST":
            content = request.form.get("content", "")
            if not content.strip():
                flash("内容不能为空", "error")
                return redirect(request.url)
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(content)
            if current_app.qdrant:
                try:
                    fm, _ = render_markdown(content)
                    current_app.qdrant.upsert_page(
                        repo_id=repo.id,
                        filename=f"{page_slug}.md",
                        title=fm.get("title", page_slug),
                        page_type=fm.get("type", "unknown"),
                        content=content,
                    )
                    current_app.qdrant.upsert_page_chunks(
                        repo_id=repo.id,
                        filename=f"{page_slug}.md",
                        title=fm.get("title", page_slug),
                        page_type=fm.get("type", "unknown"),
                        content=content,
                    )
                except Exception:
                    logger.warning("Qdrant upsert failed for edited page %s", page_slug)
            _sync_repo_counts(repo, username)
            flash("页面已保存", "success")
            return redirect(
                url_for("wiki.view_page", username=username, repo_slug=repo_slug, page_slug=page_slug)
            )

        with open(filepath, "r", encoding="utf-8") as f:
            raw = f.read()
        return render_template(
            "wiki/edit.html",
            username=username,
            repo=repo,
            page_slug=page_slug,
            content=raw,
        )

    @wiki_bp.route("/<username>/<repo_slug>/wiki/<page_slug>/delete", methods=["POST"])
    @login_required
    def delete_page(username, repo_slug, page_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        wiki_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "wiki"
        )
        filepath = os.path.join(wiki_dir, f"{page_slug}.md")
        if os.path.isfile(filepath):
            os.remove(filepath)
            if current_app.qdrant:
                try:
                    current_app.qdrant.delete_page(repo.id, f"{page_slug}.md")
                    current_app.qdrant.delete_page_chunks(repo.id, f"{page_slug}.md")
                except Exception:
                    pass
            _sync_repo_counts(repo, username)
            flash(f"页面 {page_slug} 已删除", "success")
        else:
            flash("页面不存在", "error")
        return redirect(
            url_for("repo.dashboard", username=username, repo_slug=repo_slug)
        )

    @wiki_bp.route("/<username>/<repo_slug>/wiki/search")
    def search(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return redirect(url_for("auth.login", next=request.url))
            abort(403)
        query_text = request.args.get("q", "").strip()
        results = []
        if query_text:
            wiki_dir = os.path.join(
                get_repo_path(Config.DATA_DIR, username, repo_slug), "wiki"
            )
            if os.path.isdir(wiki_dir):
                query_lower = query_text.lower()
                for filename in sorted(os.listdir(wiki_dir)):
                    if not filename.endswith(".md"):
                        continue
                    filepath = os.path.join(wiki_dir, filename)
                    try:
                        with open(filepath, "r", encoding="utf-8") as f:
                            content = f.read()
                    except OSError:
                        continue
                    content_lower = content.lower()
                    if query_lower not in content_lower:
                        continue
                    fm, _ = render_markdown(content)
                    idx = content_lower.find(query_lower)
                    start = max(0, idx - 100)
                    end = min(len(content), idx + len(query_text) + 100)
                    snippet = content[start:end].replace("\n", " ")
                    if start > 0:
                        snippet = "…" + snippet
                    if end < len(content):
                        snippet += "…"
                    results.append({
                        "slug": filename.replace(".md", ""),
                        "title": fm.get("title", filename.replace(".md", "")),
                        "type": fm.get("type", "unknown"),
                        "snippet": snippet,
                        "match_count": content_lower.count(query_lower),
                    })
            results.sort(key=lambda r: r["match_count"], reverse=True)

        return render_template(
            "wiki/search.html",
            username=username,
            repo=repo,
            query=query_text,
            results=results,
            is_owner=_is_owner(repo),
        )

    @wiki_bp.route("/<username>/<repo_slug>/wiki/export.zip")
    @login_required
    def export_zip(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)
        import io
        import zipfile as _zipfile
        wiki_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "wiki"
        )
        buf = io.BytesIO()
        with _zipfile.ZipFile(buf, "w", _zipfile.ZIP_DEFLATED) as zf:
            if os.path.isdir(wiki_dir):
                for fname in sorted(os.listdir(wiki_dir)):
                    if fname.endswith(".md"):
                        fpath = os.path.join(wiki_dir, fname)
                        zf.write(fpath, fname)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{repo_slug}-wiki.zip"',
            },
        )

    @wiki_bp.route("/<username>/<repo_slug>/search/semantic")
    @login_required
    def semantic_search(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _ensure_repo_access(repo)
        q = request.args.get("q", "").strip()
        results = []
        if q and current_app.qdrant:
            try:
                results = current_app.qdrant.search_chunks(repo_id=repo.id, query=q, limit=15)
            except Exception as exc:
                flash(f"语义检索失败：{exc}", "error")
        return render_template("wiki/semantic_search.html", username=username, repo=repo,
                               q=q, results=results)

    app.register_blueprint(wiki_bp)

    # ── Source ────────────────────────────────────────────────────────

    source_bp = Blueprint("source", __name__)

    @source_bp.route("/<username>/<repo_slug>/sources")
    def list_sources(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return redirect(url_for("auth.login", next=request.url))
            abort(403)
        raw_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "raw"
        )
        return render_template(
            "source/list.html",
            username=username,
            repo=repo,
            sources=_enrich_sources(raw_dir, repo),
            is_owner=_is_owner(repo),
            can_edit=_can_edit_repo(repo),
            can_download=current_user.is_authenticated and _can_access_repo(repo),
        )

    @source_bp.route("/<username>/<repo_slug>/sources/<source_id>")
    def view_source(username, repo_slug, source_id):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return redirect(url_for("auth.login", next=request.url))
            abort(403)
        raw_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "raw"
        )
        filepath = os.path.join(raw_dir, source_id)
        if not os.path.isfile(filepath):
            abort(404)

        with open(filepath, "r", encoding="utf-8") as f:
            raw_content = f.read()
        _, html = render_markdown(raw_content)

        ingested = (
            Task.query.filter_by(
                repo_id=repo.id, type="ingest", status="done", input_data=source_id
            ).first()
            is not None
        )

        source = {"id": source_id, "filename": source_id, "ingested": ingested}
        return render_template(
            "source/view.html",
            username=username,
            repo=repo,
            source=source,
            content=html,
            is_owner=_is_owner(repo),
            can_edit=_can_edit_repo(repo),
            can_download=current_user.is_authenticated and _can_access_repo(repo),
        )

    @source_bp.route("/<username>/<repo_slug>/sources/upload", methods=["POST"])
    @login_required
    def upload(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)

        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename:
            flash("请选择要上传的文件", "error")
            return redirect(
                url_for("source.list_sources", username=username, repo_slug=repo_slug)
            )

        safe_name = secure_filename(uploaded.filename)
        if not safe_name or not _allowed_file(safe_name):
            flash("不支持的文件格式", "error")
            return redirect(
                url_for("source.list_sources", username=username, repo_slug=repo_slug)
            )

        base = get_repo_path(Config.DATA_DIR, username, repo_slug)
        raw_dir = os.path.join(base, "raw")
        os.makedirs(raw_dir, exist_ok=True)
        ext = _file_ext(safe_name)

        saved_name = None

        if ext in ("md", "txt"):
            dest_path = os.path.join(raw_dir, safe_name)
            # 检测重复：先保存到临时文件，计算 MD5
            tmp_check = dest_path + ".uploading"
            uploaded.save(tmp_check)
            new_md5 = file_md5(tmp_check)
            dup_name = None
            for existing in os.listdir(raw_dir):
                if existing.startswith(".") or existing == safe_name or existing.endswith(".uploading"):
                    continue
                ep = os.path.join(raw_dir, existing)
                if os.path.isfile(ep):
                    try:
                        if file_md5(ep) == new_md5:
                            dup_name = existing
                            break
                    except Exception:
                        pass
            if dup_name:
                os.remove(tmp_check)
                flash(f"文件内容与已有文件「{dup_name}」重复，跳过上传。", "warning")
                return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))
            os.rename(tmp_check, dest_path)
            saved_name = safe_name
        elif ext == "csv":
            try:
                file_bytes = uploaded.read()
                decoded = file_bytes.decode("utf-8-sig")
                rows = list(csv.reader(io.StringIO(decoded)))
                saved_name = _save_tabular_source(
                    base=base,
                    raw_dir=raw_dir,
                    original_name=safe_name,
                    tables=[{"name": "CSV", "rows": rows}],
                )
                _store_original_source(raw_dir, safe_name, file_bytes)
                flash(
                    f"CSV 已转换为 Markdown + Fact Records（{safe_name} → {saved_name}）",
                    "info",
                )
            except Exception as exc:
                logger.exception("CSV parse failed for %s", safe_name)
                flash(f"CSV 解析失败: {exc}", "error")
                return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))
        elif ext in EXCEL_EXTENSIONS:
            # Excel: 用 openpyxl 解析，转换为 Markdown 表格和 Fact Records 保存
            import openpyxl
            stem = safe_name.rsplit(".", 1)[0]
            try:
                file_bytes = uploaded.read()
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                tables: list[dict] = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    tables.append({"name": sheet_name, "rows": rows})
                saved_name = _save_tabular_source(
                    base=base,
                    raw_dir=raw_dir,
                    original_name=safe_name,
                    tables=tables,
                )
                _store_original_source(raw_dir, safe_name, file_bytes)
                sheet_list = ", ".join(wb.sheetnames)
                flash(
                    f"Excel 已转换为 Markdown + Fact Records（{safe_name} → {saved_name}，"
                    f"共 {len(wb.sheetnames)} 个 Sheet：{sheet_list}）",
                    "info",
                )
            except Exception as exc:
                logger.exception("Excel parse failed for %s", safe_name)
                flash(f"Excel 解析失败: {exc}", "error")
                return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))
        elif ext in MINERU_EXTENSIONS:
            temp_dir = os.path.join(base, "temp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, safe_name)
            uploaded.save(temp_path)

            # Convert .doc → .docx via LibreOffice before sending to MinerU
            parse_path = temp_path
            converted_path = None
            if ext == "doc":
                import subprocess
                try:
                    result_conv = subprocess.run(
                        ["soffice", "--headless", "--convert-to", "docx",
                         "--outdir", temp_dir, temp_path],
                        capture_output=True, timeout=60,
                    )
                    docx_path = os.path.splitext(temp_path)[0] + ".docx"
                    if result_conv.returncode == 0 and os.path.isfile(docx_path):
                        parse_path = docx_path
                        converted_path = docx_path
                    else:
                        logger.warning("LibreOffice conversion failed for %s: %s",
                                       safe_name, result_conv.stderr.decode())
                except Exception as conv_exc:
                    logger.warning("LibreOffice conversion error for %s: %s", safe_name, conv_exc)

            try:
                result = current_app.mineru.parse_file(parse_path)
                md_content = (
                    result.get("md_content")
                    or result.get("markdown")
                    or result.get("content", "")
                )
                if not md_content:
                    flash(f"MinerU 解析 {safe_name} 未返回内容", "error")
                else:
                    md_name = os.path.splitext(safe_name)[0] + ".md"
                    with open(os.path.join(raw_dir, md_name), "w", encoding="utf-8") as f:
                        f.write(md_content)
                    saved_name = md_name
                    # 保存原始文件供下载
                    originals_dir = os.path.join(raw_dir, "originals")
                    os.makedirs(originals_dir, exist_ok=True)
                    import shutil as _shutil
                    _shutil.copy2(temp_path, os.path.join(originals_dir, safe_name))
            except MineruClientError as exc:
                logger.exception("MinerU parse failed for %s", safe_name)
                flash(f"文件解析失败: {exc}", "error")
                saved_name = None
            finally:
                if os.path.isfile(temp_path):
                    os.remove(temp_path)
                if converted_path and os.path.isfile(converted_path):
                    os.remove(converted_path)
                if os.path.isdir(temp_dir) and not os.listdir(temp_dir):
                    os.rmdir(temp_dir)
        else:
            flash("不支持的文件格式", "error")

        _sync_repo_counts(repo, username)

        is_xhr = request.headers.get("X-Requested-With") == "XMLHttpRequest" or "XMLHttpRequest" in (request.headers.get("X-Requested-With", ""))
        if not is_xhr:
            is_xhr = "application/json" in request.accept_mimetypes

        if saved_name:
            task = Task(
                repo_id=repo.id,
                type="ingest",
                status="queued",
                input_data=saved_name,
            )
            db.session.add(task)
            db.session.commit()

            if is_xhr:
                return jsonify(ok=True, filename=saved_name, task_id=task.id)

            flash(
                f"文件 {saved_name} 上传成功，摄入任务已排队（#{task.id}）",
                "success",
            )
        elif is_xhr:
            return jsonify(ok=False, error="上传失败"), 400

        return redirect(
            url_for("source.list_sources", username=username, repo_slug=repo_slug)
        )

    @source_bp.route(
        "/<username>/<repo_slug>/sources/<source_id>/download", methods=["GET"]
    )
    @login_required
    def download_source(username, repo_slug, source_id):
        user, repo = _get_repo_or_404(username, repo_slug)
        _ensure_repo_access(repo)
        if not current_user.is_authenticated:
            abort(403)
        raw_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "raw"
        )
        # 优先返回原始文件（PDF/Word 等），如果存在的话
        stem = os.path.splitext(source_id)[0]
        originals_dir = os.path.join(raw_dir, "originals")
        original_file = None
        if os.path.isdir(originals_dir):
            for fname in os.listdir(originals_dir):
                if os.path.splitext(fname)[0] == stem:
                    original_file = os.path.join(originals_dir, fname)
                    break
        if original_file and os.path.isfile(original_file):
            return send_file(
                original_file,
                as_attachment=True,
                download_name=os.path.basename(original_file),
            )
        # 没有原始文件则下载 .md 文件
        filepath = os.path.join(raw_dir, source_id)
        if not os.path.isfile(filepath):
            flash("源文件不存在", "error")
            return redirect(
                url_for("source.list_sources", username=username, repo_slug=repo_slug)
            )
        return send_file(
            filepath,
            as_attachment=True,
            download_name=source_id,
        )

    @source_bp.route(
        "/<username>/<repo_slug>/sources/<source_id>/delete", methods=["POST"]
    )
    @login_required
    def delete_source(username, repo_slug, source_id):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        raw_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "raw"
        )
        filepath = os.path.join(raw_dir, source_id)
        if os.path.isfile(filepath):
            os.remove(filepath)
            removed = _purge_source_wiki(current_app._get_current_object(), repo, username, source_id)
            flash(f"已删除 {source_id}（清理了 {removed} 个关联 Wiki 页面）", "success")
        else:
            flash("文件不存在", "error")
        _sync_repo_counts(repo, username)
        return redirect(
            url_for("source.list_sources", username=username, repo_slug=repo_slug)
        )

    @source_bp.route(
        "/<username>/<repo_slug>/sources/<source_id>/rename", methods=["POST"]
    )
    @login_required
    def rename_source(username, repo_slug, source_id):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        new_name = request.form.get("new_name", "").strip()
        if not new_name:
            flash("文件名不能为空", "error")
            return redirect(
                url_for(
                    "source.list_sources", username=username, repo_slug=repo_slug
                )
            )
        safe_new = secure_filename(new_name)
        if not safe_new:
            flash("文件名包含非法字符", "error")
            return redirect(
                url_for(
                    "source.list_sources", username=username, repo_slug=repo_slug
                )
            )
        raw_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "raw"
        )
        old_path = os.path.join(raw_dir, source_id)
        new_path = os.path.join(raw_dir, safe_new)
        if not os.path.isfile(old_path):
            flash("源文件不存在", "error")
        elif os.path.exists(new_path):
            flash(f"{safe_new} 已存在", "error")
        else:
            os.rename(old_path, new_path)
            # Queue re-ingest so wiki reflects the renamed source
            task = Task(repo_id=repo.id, type="ingest", status="queued", input_data=safe_new)
            db.session.add(task)
            db.session.commit()
            flash(f"已重命名为 {safe_new}，已排队重新摄入（#{task.id}）", "success")
        return redirect(
            url_for("source.list_sources", username=username, repo_slug=repo_slug)
        )

    @source_bp.route(
        "/<username>/<repo_slug>/sources/batch-delete", methods=["POST"]
    )
    @login_required
    def batch_delete(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        filenames = request.form.getlist("files")
        raw_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "raw"
        )
        deleted = 0
        purged_wiki = 0
        app_obj = current_app._get_current_object()
        for fn in filenames:
            fp = os.path.join(raw_dir, secure_filename(fn))
            if os.path.isfile(fp):
                os.remove(fp)
                purged_wiki += _purge_source_wiki(app_obj, repo, username, secure_filename(fn))
                deleted += 1
        _sync_repo_counts(repo, username)
        flash(f"已删除 {deleted} 个文件，清理了 {purged_wiki} 个关联 Wiki 页面", "success")
        return redirect(
            url_for("source.list_sources", username=username, repo_slug=repo_slug)
        )

    @source_bp.route("/<username>/<repo_slug>/sources/batch-ingest", methods=["POST"])
    @login_required
    def batch_ingest(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        raw_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "raw"
        )
        selected = request.form.getlist("files")
        ingested_files = {
            t.input_data
            for t in Task.query.filter_by(repo_id=repo.id, type="ingest", status="done").all()
            if t.input_data
        }
        queued_files = {
            t.input_data
            for t in Task.query.filter(
                Task.repo_id == repo.id,
                Task.type == "ingest",
                Task.status.in_(["queued", "running"]),
            ).all()
            if t.input_data
        }
        all_sources = list_raw_sources(raw_dir)
        queued_count = 0
        for src in all_sources:
            fn = src["filename"]
            if selected and fn not in selected:
                continue
            if fn in ingested_files or fn in queued_files:
                continue
            task = Task(repo_id=repo.id, type="ingest", status="queued", input_data=fn)
            db.session.add(task)
            queued_count += 1
        if queued_count:
            db.session.commit()
        flash(f"已排队 {queued_count} 个摄入任务", "success" if queued_count else "info")
        return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))

    @source_bp.route("/<username>/<repo_slug>/sources/import-url", methods=["POST"])
    @login_required
    def import_url(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        url = request.form.get("url", "").strip()
        if not url:
            flash("URL 不能为空", "error")
            return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))

        if not url.startswith(("http://", "https://")):
            flash("请输入有效的 HTTP/HTTPS URL", "error")
            return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))

        try:
            import trafilatura
            downloaded = trafilatura.fetch_url(url)
            if not downloaded:
                flash("无法获取页面内容，请检查 URL 是否可访问", "error")
                return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))
            text = trafilatura.extract(
                downloaded,
                output_format="markdown",
                include_links=True,
                include_images=False,
                no_fallback=False,
            )
            if not text:
                flash("无法提取页面正文内容", "error")
                return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))
        except Exception as exc:
            logger.exception("URL import failed for %s", url)
            flash(f"导入失败: {exc}", "error")
            return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))

        from urllib.parse import urlparse
        parsed = urlparse(url)
        raw_slug = slugify(parsed.netloc + parsed.path)
        path_slug = (raw_slug or "imported")[:60]
        filename = f"{path_slug}.md"

        raw_dir = os.path.join(get_repo_path(Config.DATA_DIR, username, repo_slug), "raw")
        os.makedirs(raw_dir, exist_ok=True)

        save_name = filename
        counter = 0
        while os.path.exists(os.path.join(raw_dir, save_name)):
            counter += 1
            save_name = f"{path_slug}-{counter}.md"

        header = f"---\ntitle: {url}\nsource_url: {url}\n---\n\n"
        with open(os.path.join(raw_dir, save_name), "w", encoding="utf-8") as f:
            f.write(header + text)

        task = Task(repo_id=repo.id, type="ingest", status="queued", input_data=save_name)
        db.session.add(task)
        db.session.commit()
        _sync_repo_counts(repo, username)
        flash(f"已导入 {save_name}，摄入任务已排队（#{task.id}）", "success")
        return redirect(url_for("source.list_sources", username=username, repo_slug=repo_slug))

    @source_bp.route("/<username>/<repo_slug>/ingest/<source_id>", methods=["POST"])
    @login_required
    def ingest(username, repo_slug, source_id):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)

        raw_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "raw"
        )
        if not os.path.isfile(os.path.join(raw_dir, source_id)):
            flash("源文件不存在", "error")
            return redirect(
                url_for("source.list_sources", username=username, repo_slug=repo_slug)
            )

        task = Task(repo_id=repo.id, type="ingest", status="queued", input_data=source_id)
        db.session.add(task)
        db.session.commit()

        flash(f"摄入任务已排队（#{task.id}）", "success")
        return redirect(
            url_for("ops.task_queue", username=username, repo_slug=repo_slug)
        )

    app.register_blueprint(source_bp)

    # ── Ops ───────────────────────────────────────────────────────────

    ops_bp = Blueprint("ops", __name__)

    @ops_bp.route("/<username>/<repo_slug>/ingest-progress/<int:task_id>")
    @login_required
    def ingest_progress(username, repo_slug, task_id):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        task = Task.query.get_or_404(task_id)
        if task.repo_id != repo.id:
            abort(404)
        return render_template(
            "ops/ingest_progress.html",
            username=username,
            repo=repo,
            task_id=task_id,
        )

    @ops_bp.route("/task/<int:task_id>/stream")
    @login_required
    def task_stream(task_id):
        """SSE endpoint that polls task progress from DB (worker runs in background)."""
        task = Task.query.get_or_404(task_id)
        repo = Repo.query.get_or_404(task.repo_id)
        _require_editor(repo)

        import time

        last_msg = ""

        def generate():
            nonlocal last_msg
            while True:
                db.session.expire_all()
                t = db.session.get(Task, task_id)
                if not t:
                    yield _sse("error_event", message="Task not found")
                    return

                msg = t.progress_msg or ""
                pct = t.progress or 0

                if msg != last_msg:
                    yield _sse("progress", percent=pct, message=msg)
                    yield _sse("step", message=msg, status="running")
                    last_msg = msg

                if t.status == "done":
                    yield _sse("progress", percent=100, message=msg)
                    yield _sse("done", message=msg)
                    return
                elif t.status == "failed":
                    yield _sse("error_event", message=msg or "Task failed")
                    return
                elif t.status == "cancelled":
                    yield _sse("error_event", message=msg or "Task cancelled")
                    return

                time.sleep(1)

        return Response(
            stream_with_context(generate()),
            mimetype="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    @ops_bp.route("/<username>/<repo_slug>/tasks")
    @login_required
    def task_queue(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)
        tasks = (
            Task.query.filter_by(repo_id=repo.id)
            .order_by(Task.created_at.desc())
            .limit(50)
            .all()
        )
        return render_template(
            "ops/tasks.html",
            username=username,
            repo=repo,
            tasks=tasks,
            is_owner=_is_owner(repo),
            can_edit=_can_edit_repo(repo),
        )

    @ops_bp.route("/api/tasks/<int:task_id>/status")
    @login_required
    def task_status_api(task_id):
        task = Task.query.get_or_404(task_id)
        repo = Repo.query.get_or_404(task.repo_id)
        _require_editor(repo)
        return jsonify(
            id=task.id,
            status=task.status,
            progress=task.progress,
            progress_msg=task.progress_msg,
            input_data=task.input_data,
            cancel_requested=task.cancel_requested,
        )

    @ops_bp.route("/api/tasks/<int:task_id>/retry", methods=["POST"])
    @login_required
    def retry_task(task_id):
        task = Task.query.get_or_404(task_id)
        repo = db.session.get(Repo, task.repo_id)
        if repo is None:
            abort(404)
        _require_editor(repo)
        if task.status not in ("failed", "cancelled"):
            return jsonify(error="只能重试失败或已取消的任务"), 400
        task.status = "queued"
        task.progress = 0
        task.progress_msg = ""
        task.cancel_requested = False
        task.started_at = None
        task.finished_at = None
        db.session.commit()
        return jsonify(ok=True, task_id=task.id)

    @ops_bp.route("/api/tasks/<int:task_id>/cancel", methods=["POST"])
    @login_required
    def cancel_task(task_id):
        task = Task.query.get_or_404(task_id)
        repo = db.session.get(Repo, task.repo_id)
        if repo is None:
            abort(404)
        _require_editor(repo)
        if task.status == "queued":
            task.status = "cancelled"
            task.cancel_requested = True
            task.progress_msg = "Task cancelled before start"
            task.finished_at = datetime.now(timezone.utc)
            db.session.commit()
            return jsonify(ok=True, task_id=task.id, status=task.status)
        if task.status == "running":
            task.cancel_requested = True
            task.progress_msg = "已请求终止，等待当前步骤结束"
            db.session.commit()
            return jsonify(ok=True, task_id=task.id, status=task.status, cancel_requested=True)
        return jsonify(error="当前任务无法取消"), 400

    @ops_bp.route(
        "/<username>/<repo_slug>/query", methods=["GET"], endpoint="query"
    )
    def query_page(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        access_result = _ensure_repo_access(repo)
        if access_result is not None:
            return access_result
        return render_template(
            "ops/query.html",
            username=username,
            repo=repo,
            can_edit=_can_edit_repo(repo),
        )

    @ops_bp.route(
        "/<username>/<repo_slug>/query", methods=["POST"], endpoint="query_api"
    )
    def query_api(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return jsonify(error="请先登录"), 401
            abort(403)
        data = request.get_json(silent=True) or {}
        question = data.get("q", "").strip()
        session_key = data.get("session_key", "")
        history: list[dict] = []
        if session_key and current_user.is_authenticated:
            from models import ConversationSession
            cs = ConversationSession.query.filter_by(
                repo_id=repo.id, user_id=current_user.id, session_key=session_key
            ).first()
            if cs:
                try:
                    history = json.loads(cs.messages_json)
                except Exception:
                    history = []

        def _persist_session(answer_markdown: str) -> None:
            if not session_key or not current_user.is_authenticated:
                return
            try:
                session_history = history + [
                    {"role": "user", "content": question},
                    {"role": "assistant", "content": answer_markdown},
                ]
                session_history = session_history[-20:]
                from models import ConversationSession
                cs = ConversationSession.query.filter_by(
                    repo_id=repo.id, user_id=current_user.id, session_key=session_key
                ).first()
                if cs:
                    cs.messages_json = json.dumps(session_history, ensure_ascii=False)
                    if cs.title == "新对话":
                        cs.title = question[:60]
                else:
                    cs = ConversationSession(
                        repo_id=repo.id,
                        user_id=current_user.id,
                        session_key=session_key,
                        title=question[:60],
                        messages_json=json.dumps(session_history, ensure_ascii=False),
                    )
                    db.session.add(cs)
                db.session.commit()
            except Exception as sess_exc:
                logger.warning("Session update failed: %s", sess_exc)

        pre_answer = data.get("_rendered_answer")
        if pre_answer is not None:
            _, answer_html = render_markdown(pre_answer, _wiki_base_url(username, repo_slug))
            pre_wiki = data.get("_wiki_sources", [])
            pre_qdrant = data.get("_qdrant_sources", [])
            pre_confidence = data.get("_confidence", {"level": "low", "score": 0.0, "reasons": []})
            pre_wiki_ev = data.get("_wiki_evidence", []) or []
            pre_chunk_ev = data.get("_chunk_evidence", []) or []
            pre_fact_ev = data.get("_fact_evidence", []) or []
            pre_ev_summary = data.get("_evidence_summary", "")

            import uuid as _uuid
            import json as _json
            _pre_trace_id = _uuid.uuid4().hex[:8]
            try:
                from models import QueryLog
                _pre_conf = pre_confidence if isinstance(pre_confidence, dict) else {}
                _pre_retrieval = {
                    "wiki": [{"filename": e.get("filename", ""), "title": e.get("title", ""), "reason": e.get("reason", "")} for e in pre_wiki_ev],
                    "chunks": [{"filename": e.get("filename", ""), "score": e.get("score"), "snippet": (e.get("snippet") or "")[:300]} for e in pre_chunk_ev],
                    "facts": [{"source_file": e.get("source_file", ""), "score": e.get("score"), "fields": e.get("fields", {})} for e in pre_fact_ev],
                }
                _pre_ql = QueryLog(
                    trace_id=_pre_trace_id,
                    repo_id=repo.id,
                    user_id=current_user.id if current_user.is_authenticated else None,
                    question=question,
                    answer_preview=pre_answer[:500],
                    full_answer=pre_answer,
                    confidence=_pre_conf.get("level", "low"),
                    wiki_hit_count=len(pre_wiki_ev),
                    chunk_hit_count=len(pre_chunk_ev),
                    used_wiki_pages=_json.dumps([e.get("filename", "") for e in pre_wiki_ev], ensure_ascii=False),
                    used_chunk_ids=_json.dumps([e.get("chunk_id", "") for e in pre_chunk_ev], ensure_ascii=False),
                    evidence_summary=pre_ev_summary,
                    retrieval_json=_json.dumps(_pre_retrieval, ensure_ascii=False),
                    query_mode="stream",
                )
                db.session.add(_pre_ql)
                db.session.commit()
            except Exception as _pre_ql_exc:
                logger.warning("QueryLog (stream) write failed: %s", _pre_ql_exc)

            def _src_to_ref(fn):
                slug = fn.replace(".md", "")
                return {
                    "url": url_for("wiki.view_page", username=username,
                                   repo_slug=repo_slug, page_slug=slug),
                    "title": fn.replace(".md", "").replace("-", " ").title(),
                    "filename": fn,
                }

            _persist_session(pre_answer)
            return jsonify(
                html=answer_html,
                markdown=pre_answer,
                answer=pre_answer,
                trace_id=_pre_trace_id,
                confidence=pre_confidence,
                wiki_evidence=pre_wiki_ev,
                chunk_evidence=pre_chunk_ev,
                fact_evidence=pre_fact_ev,
                evidence_summary=pre_ev_summary,
                references=[_src_to_ref(fn) for fn in pre_wiki],
                wiki_sources=[_src_to_ref(fn) for fn in pre_wiki],
                qdrant_sources=[_src_to_ref(fn) for fn in pre_qdrant],
            )

        if not question:
            return jsonify(error="请输入问题"), 400

        import time as _time
        _t0 = _time.monotonic()
        try:
            result = current_app.wiki_engine.query_with_evidence(
                repo, username, question, _wiki_base_url(username, repo_slug),
                history=history[-6:] if history else None,
            )
        except Exception as exc:
            logger.exception("Query failed for repo %s", repo.id)
            return jsonify(error=f"查询失败: {exc}"), 500
        latency_ms = int((_time.monotonic() - _t0) * 1000)

        _, answer_html = render_markdown(
            result.get("markdown", ""), _wiki_base_url(username, repo_slug)
        )

        # -- Write query log (DB + file) -----------------------------------
        _wiki_ev  = result.get("wiki_evidence", []) or []
        _chunk_ev = result.get("chunk_evidence", []) or []
        _fact_ev  = result.get("fact_evidence", []) or []
        _conf     = result.get("confidence", {}) or {}
        _mode     = result.get("query_mode", "")
        _answer   = result.get("markdown", "")
        _username = current_user.username if current_user.is_authenticated else None

        import uuid as _uuid
        _trace_id = _uuid.uuid4().hex[:8]  # 8-char hex, e.g. "a3f9c12b"

        try:
            from models import QueryLog
            import json as _json
            _retrieval = {
                "wiki": [
                    {"filename": e.get("filename", ""), "title": e.get("title", ""), "reason": e.get("reason", "")}
                    for e in _wiki_ev
                ],
                "chunks": [
                    {"filename": e.get("filename", ""), "score": e.get("score"), "snippet": (e.get("snippet") or "")[:300]}
                    for e in _chunk_ev
                ],
                "facts": [
                    {"source_file": e.get("source_file", ""), "score": e.get("score"), "fields": e.get("fields", {})}
                    for e in _fact_ev
                ],
            }
            ql = QueryLog(
                trace_id=_trace_id,
                repo_id=repo.id,
                user_id=current_user.id if current_user.is_authenticated else None,
                question=question,
                answer_preview=_answer[:500],
                full_answer=_answer,
                confidence=_conf.get("level", "low"),
                wiki_hit_count=len(_wiki_ev),
                chunk_hit_count=len(_chunk_ev),
                used_wiki_pages=_json.dumps([e.get("filename", "") for e in _wiki_ev], ensure_ascii=False),
                used_chunk_ids=_json.dumps([e.get("chunk_id", "") for e in _chunk_ev], ensure_ascii=False),
                evidence_summary=result.get("evidence_summary", ""),
                retrieval_json=_json.dumps(_retrieval, ensure_ascii=False),
                query_mode=_mode,
                latency_ms=latency_ms,
            )
            db.session.add(ql)
            db.session.commit()
        except Exception as ql_exc:
            logger.warning("QueryLog write failed: %s", ql_exc)

        try:
            current_app.query_trace_logger.write(
                repo=f"{username}/{repo.slug}",
                user=_username,
                question=question,
                mode=_mode,
                latency_ms=latency_ms,
                confidence=_conf,
                wiki_evidence=_wiki_ev,
                chunk_evidence=_chunk_ev,
                fact_evidence=_fact_ev,
                answer=_answer,
                trace_id=_trace_id,
            )
        except Exception as tl_exc:
            logger.warning("QueryTraceLogger write failed: %s", tl_exc)

        # -- Update conversation session ----------------------------------
        _persist_session(result.get("markdown", ""))

        def _fn_to_ref(fn: str) -> dict:
            slug = fn.replace(".md", "")
            return {
                "url": url_for("wiki.view_page", username=username,
                               repo_slug=repo_slug, page_slug=slug),
                "title": fn.replace(".md", "").replace("-", " ").title(),
                "filename": fn,
            }

        wiki_sources_refs = [_fn_to_ref(fn) for fn in result.get("wiki_sources", [])]
        qdrant_sources_refs = [_fn_to_ref(fn) for fn in result.get("qdrant_sources", [])]
        referenced = result.get("referenced_pages", [])
        references = [_fn_to_ref(fn) for fn in referenced]

        return jsonify(
            html=answer_html,
            markdown=result.get("markdown", ""),
            answer=result.get("markdown", ""),
            trace_id=_trace_id,
            confidence=result.get("confidence", {}),
            wiki_evidence=result.get("wiki_evidence", []),
            chunk_evidence=result.get("chunk_evidence", []),
            fact_evidence=result.get("fact_evidence", []),
            evidence_summary=result.get("evidence_summary", ""),
            referenced_pages=referenced,
            references=references,
            wiki_sources=wiki_sources_refs,
            qdrant_sources=qdrant_sources_refs,
        )

    @ops_bp.route("/<username>/<repo_slug>/query/stream", methods=["GET"])
    def query_stream(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return jsonify(error="请先登录"), 401
            abort(403)
        question = request.args.get("q", "").strip()
        if not question:
            return jsonify(error="请输入问题"), 400

        def generate():
            try:
                for event_dict in current_app.wiki_engine.query_stream(repo, username, question):
                    event = event_dict["event"]
                    data = json.dumps(event_dict["data"], ensure_ascii=False)
                    yield f"event: {event}\ndata: {data}\n\n"
            except Exception as exc:
                logger.exception("query_stream error for repo %s", repo.id)
                yield f"event: error\ndata: {json.dumps({'message': str(exc)})}\n\n"

        return Response(
            stream_with_context(generate()),
            mimetype="text/event-stream",
            headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
        )

    @ops_bp.route("/<username>/<repo_slug>/query/save", methods=["POST"])
    @login_required
    def save_query_page(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_editor(repo)

        content = request.form.get("content", "")
        query_text = request.form.get("query", "").strip()
        if not content:
            flash("没有可保存的内容", "error")
            return redirect(
                url_for("ops.query", username=username, repo_slug=repo_slug)
            )

        slug = slugify(query_text) if query_text else "query-result"
        slug = slug or "query-result"
        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

        if not content.startswith("---"):
            content = (
                f"---\ntitle: {query_text or 'Query Result'}\n"
                f"type: reference\nupdated: {now_str}\n---\n\n{content}"
            )

        wiki_dir = os.path.join(
            get_repo_path(Config.DATA_DIR, username, repo_slug), "wiki"
        )
        os.makedirs(wiki_dir, exist_ok=True)

        filename = f"{slug}.md"
        counter = 1
        while os.path.exists(os.path.join(wiki_dir, filename)):
            filename = f"{slug}-{counter}.md"
            counter += 1

        with open(os.path.join(wiki_dir, filename), "w", encoding="utf-8") as f:
            f.write(content)

        if current_app.qdrant:
            try:
                fm, _ = render_markdown(content)
                current_app.qdrant.upsert_page(
                    repo_id=repo.id,
                    filename=filename,
                    title=fm.get("title", slug),
                    page_type=fm.get("type", "reference"),
                    content=content,
                )
            except Exception:
                logger.warning(
                    "Qdrant upsert failed for saved query page %s", filename
                )

        _sync_repo_counts(repo, username)
        flash(f"已保存为 Wiki 页面: {filename}", "success")
        return redirect(
            url_for(
                "wiki.view_page",
                username=username,
                repo_slug=repo_slug,
                page_slug=filename.replace(".md", ""),
            )
        )

    @ops_bp.route("/<username>/<repo_slug>/lint", methods=["GET", "POST"])
    @login_required
    def lint(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)

        try:
            raw_result = current_app.wiki_engine.lint(repo, username)
        except Exception as exc:
            logger.exception("Lint failed for repo %s", repo.id)
            flash(f"检查失败: {exc}", "error")
            return redirect(
                url_for("repo.dashboard", username=username, repo_slug=repo_slug)
            )

        issue_bucket_map = {
            "contradiction": "contradictions",
            "orphan": "orphan_pages",
            "missing_link": "missing_refs",
            "bad_frontmatter": "missing_refs",
            "wrong_type": "missing_refs",
        }
        report: dict[str, list] = {
            "contradictions": [],
            "stale_claims": [],
            "orphan_pages": [],
            "missing_pages": [],
            "missing_refs": [],
            "suggestions": [],
        }

        for issue in raw_result.get("issues", []):
            bucket = issue_bucket_map.get(issue.get("type", ""), "missing_refs")
            page_file = issue.get("page", "")
            report[bucket].append(
                {
                    "page_slug": page_file.replace(".md", "") if page_file else None,
                    "title": page_file.replace(".md", "") if page_file else None,
                    "message": issue.get("description", ""),
                }
            )

        for suggestion in raw_result.get("suggestions", []):
            report["suggestions"].append(
                {"page_slug": None, "title": None, "message": suggestion}
            )

        fixable_types = {"bad_frontmatter", "orphan", "missing_link", "wrong_type"}
        raw_issues = raw_result.get("issues", [])
        has_fixes = any(i.get("type") in fixable_types for i in raw_issues)

        return render_template(
            "ops/lint.html",
            username=username,
            repo=repo,
            report=report,
            has_fixes=has_fixes,
            raw_issues=raw_issues,
        )

    @ops_bp.route("/<username>/<repo_slug>/apply-fixes", methods=["POST"])
    @login_required
    def apply_fixes(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)

        issues_json = request.form.get("issues_json", "[]")
        try:
            issues = json.loads(issues_json)
        except json.JSONDecodeError:
            issues = []

        if not issues:
            flash("没有可修复的问题", "info")
            return redirect(url_for("ops.lint", username=username, repo_slug=repo_slug))

        try:
            result = current_app.wiki_engine.apply_fixes(repo, username, issues)
        except Exception as exc:
            logger.exception("apply_fixes failed for repo %s", repo.id)
            flash(f"修复失败: {exc}", "error")
            return redirect(url_for("ops.lint", username=username, repo_slug=repo_slug))

        fixed_count = len(result.get("fixed", []))
        skipped_count = len(result.get("skipped", []))
        flash(
            f"已修复 {fixed_count} 个问题，跳过 {skipped_count} 个（矛盾类问题需人工审查）",
            "success" if fixed_count > 0 else "info",
        )
        return redirect(url_for("ops.lint", username=username, repo_slug=repo_slug))

    @ops_bp.route("/<username>/<repo_slug>/session", methods=["GET"])
    @login_required
    def get_session(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_manage_sessions(repo):
            abort(403)
        key = request.args.get("key", "")
        if not key:
            return jsonify(messages=[])
        from models import ConversationSession
        cs = ConversationSession.query.filter_by(
            repo_id=repo.id, user_id=current_user.id, session_key=key
        ).first()
        msgs = json.loads(cs.messages_json) if cs else []
        return jsonify(messages=msgs)

    @ops_bp.route("/<username>/<repo_slug>/render-markdown", methods=["POST"])
    def render_markdown_api(username, repo_slug):
        """轻量 Markdown 渲染接口，供前端历史消息回显使用。"""
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_access_repo(repo):
            if not current_user.is_authenticated:
                return jsonify(error="请先登录"), 401
            abort(403)
        data = request.get_json(silent=True) or {}
        base_url = _wiki_base_url(username, repo_slug)
        # batch mode: {"messages": ["md1", "md2", ...]}
        if "messages" in data:
            html_list = []
            for raw in data["messages"]:
                _, html = render_markdown(raw or "", base_url)
                html_list.append(html)
            return jsonify(html_list=html_list)
        # single mode: {"markdown": "..."}
        raw = data.get("markdown", "")
        _, html = render_markdown(raw, base_url)
        return jsonify(html=html)

    @ops_bp.route("/<username>/<repo_slug>/sessions", methods=["GET"])
    @login_required
    def list_sessions(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_manage_sessions(repo):
            abort(403)
        from models import ConversationSession
        sessions = (
            ConversationSession.query
            .filter_by(repo_id=repo.id, user_id=current_user.id)
            .order_by(ConversationSession.updated_at.desc())
            .limit(50)
            .all()
        )
        return jsonify(sessions=[{
            "key": s.session_key,
            "title": s.title,
            "updated_at": s.updated_at.strftime("%Y-%m-%d %H:%M"),
            "message_count": len(json.loads(s.messages_json or "[]")) // 2,
        } for s in sessions])

    @ops_bp.route("/<username>/<repo_slug>/sessions/new", methods=["POST"])
    @login_required
    def new_session(username, repo_slug):
        import uuid
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_manage_sessions(repo):
            abort(403)
        from models import ConversationSession
        key = "sess_" + uuid.uuid4().hex[:16]
        cs = ConversationSession(
            repo_id=repo.id, user_id=current_user.id,
            session_key=key, title="新对话", messages_json="[]"
        )
        db.session.add(cs)
        db.session.commit()
        return jsonify(ok=True, key=key, title="新对话")

    @ops_bp.route("/<username>/<repo_slug>/sessions/<session_key>/delete", methods=["POST"])
    @login_required
    def delete_session(username, repo_slug, session_key):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_manage_sessions(repo):
            abort(403)
        from models import ConversationSession
        cs = ConversationSession.query.filter_by(
            repo_id=repo.id, user_id=current_user.id, session_key=session_key
        ).first()
        if cs:
            db.session.delete(cs)
            db.session.commit()
        return jsonify(ok=True)

    @ops_bp.route("/<username>/<repo_slug>/sessions/<session_key>/rename", methods=["POST"])
    @login_required
    def rename_session(username, repo_slug, session_key):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_manage_sessions(repo):
            abort(403)
        title = (request.get_json(silent=True) or {}).get("title", "").strip()
        if not title:
            return jsonify(error="标题不能为空"), 400
        from models import ConversationSession
        cs = ConversationSession.query.filter_by(
            repo_id=repo.id, user_id=current_user.id, session_key=session_key
        ).first()
        if cs:
            cs.title = title[:100]
            db.session.commit()
        return jsonify(ok=True)

    @ops_bp.route("/<username>/<repo_slug>/session/clear", methods=["POST"])
    @login_required
    def clear_session(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        if not _can_manage_sessions(repo):
            abort(403)
        key = (request.get_json(silent=True) or {}).get("key", "")
        if key:
            from models import ConversationSession
            cs = ConversationSession.query.filter_by(
                repo_id=repo.id, user_id=current_user.id, session_key=key
            ).first()
            if cs:
                cs.messages_json = "[]"
                db.session.commit()
        return jsonify(ok=True)

    @ops_bp.route("/<username>/<repo_slug>/insights")
    @login_required
    def insights(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)
        from models import QueryLog
        low_conf_logs = (
            QueryLog.query.filter_by(repo_id=repo.id, confidence="low")
            .order_by(QueryLog.created_at.desc())
            .limit(50)
            .all()
        )
        gaps = None
        if request.args.get("analyze"):
            query_log_data = [{"question": q.question, "confidence": q.confidence} for q in low_conf_logs]
            try:
                gaps = current_app.wiki_engine.find_gaps(repo, username, query_log_data)
            except Exception as exc:
                flash(f"分析失败：{exc}", "error")
        return render_template("ops/insights.html", username=username, repo=repo,
                               low_conf_logs=low_conf_logs, gaps=gaps)

    @ops_bp.route("/<username>/<repo_slug>/entity-check")
    @login_required
    def entity_check(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        _require_owner(repo)
        result = None
        if request.args.get("analyze"):
            try:
                result = current_app.wiki_engine.find_entity_duplicates(repo, username)
            except Exception as exc:
                flash(f"检测失败：{exc}", "error")
        return render_template("ops/entity_check.html", username=username, repo=repo, result=result)

    @ops_bp.route("/<username>/<repo_slug>/feedback", methods=["POST"])
    def submit_feedback(username, repo_slug):
        user, repo = _get_repo_or_404(username, repo_slug)
        data = request.get_json(silent=True) or {}
        trace_id = data.get("trace_id", "").strip()
        rating = data.get("rating", "").strip()
        comment = data.get("comment", "").strip() or None
        if not trace_id or rating not in ("good", "bad"):
            return jsonify(error="参数错误"), 400
        from models import QueryFeedback
        fb = QueryFeedback(
            trace_id=trace_id,
            repo_id=repo.id,
            user_id=current_user.id if current_user.is_authenticated else None,
            rating=rating,
            comment=comment,
        )
        db.session.add(fb)
        db.session.commit()
        return jsonify(ok=True)

    app.register_blueprint(ops_bp)

    # ── Admin ─────────────────────────────────────────────────────────

    admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

    def _require_admin():
        if not current_user.is_authenticated or current_user.username != Config.ADMIN_USERNAME:
            abort(403)

    @admin_bp.route("/")
    @login_required
    def dashboard():
        _require_admin()

        user_count = User.query.count()
        repo_count = Repo.query.count()
        task_total = Task.query.count()
        task_done = Task.query.filter_by(status="done").count()
        task_failed = Task.query.filter_by(status="failed").count()
        task_running = Task.query.filter(
            Task.status.in_(["queued", "running"])
        ).count()

        data_dir = Config.DATA_DIR
        disk_bytes = 0
        if os.path.isdir(data_dir):
            try:
                for dirpath, _dirnames, filenames in os.walk(data_dir):
                    for fname in filenames:
                        try:
                            disk_bytes += os.path.getsize(os.path.join(dirpath, fname))
                        except OSError:
                            pass
            except Exception:
                pass

        recent_users = (
            User.query.order_by(User.created_at.desc()).limit(10).all()
        )

        return render_template(
            "admin/dashboard.html",
            user_count=user_count,
            repo_count=repo_count,
            task_total=task_total,
            task_done=task_done,
            task_failed=task_failed,
            task_running=task_running,
            disk_bytes=disk_bytes,
            recent_users=recent_users,
        )

    @admin_bp.route("/audit")
    @login_required
    def audit_log():
        _require_admin()
        page = request.args.get("page", 1, type=int)
        from models import AuditLog
        logs = AuditLog.query.order_by(AuditLog.created_at.desc()).paginate(page=page, per_page=50)
        return render_template("admin/audit.html", logs=logs)

    @admin_bp.route("/health")
    @login_required
    def health_detail():
        _require_admin()
        import httpx as _httpx
        checks: dict = {}
        services = {
            "qdrant": Config.QDRANT_URL + "/healthz",
            "embedding": Config.EMBEDDING_API_BASE.rstrip("/") + "/models",
            "mineru": Config.MINERU_API_URL + "/health",
        }
        for name, url in services.items():
            try:
                r = _httpx.get(url, timeout=5)
                checks[name] = {
                    "status": "ok" if r.status_code < 400 else "error",
                    "latency_ms": round(r.elapsed.total_seconds() * 1000),
                }
            except Exception as exc:
                checks[name] = {"status": "error", "error": str(exc)[:80]}
        checks["task_queue"] = {
            "queued": Task.query.filter_by(status="queued").count(),
            "running": Task.query.filter_by(status="running").count(),
            "failed": Task.query.filter_by(status="failed").count(),
        }
        return render_template("admin/health.html", checks=checks)

    @admin_bp.route("/query-stats")
    @login_required
    def query_stats():
        _require_admin()
        from models import QueryLog
        from sqlalchemy import func
        total = QueryLog.query.count()
        by_conf = db.session.query(
            QueryLog.confidence, func.count(QueryLog.id)
        ).group_by(QueryLog.confidence).all()
        by_repo = (
            db.session.query(Repo.name, func.count(QueryLog.id))
            .join(QueryLog, QueryLog.repo_id == Repo.id)
            .group_by(Repo.id)
            .order_by(func.count(QueryLog.id).desc())
            .limit(10)
            .all()
        )
        recent_low = (
            QueryLog.query.filter_by(confidence="low")
            .order_by(QueryLog.created_at.desc())
            .limit(20)
            .all()
        )
        return render_template("admin/query_stats.html", total=total,
                               by_conf=by_conf, by_repo=by_repo, recent_low=recent_low)

    @admin_bp.route("/query-logs")
    @login_required
    def query_logs():
        _require_admin()
        from models import QueryLog
        page = request.args.get("page", 1, type=int)
        q_filter = request.args.get("q", "").strip()
        conf_filter = request.args.get("conf", "").strip()
        user_filter = request.args.get("user", "").strip()

        query = (
            db.session.query(QueryLog, User, Repo)
            .outerjoin(User, QueryLog.user_id == User.id)
            .join(Repo, QueryLog.repo_id == Repo.id)
            .order_by(QueryLog.created_at.desc())
        )
        if q_filter:
            query = query.filter(QueryLog.question.ilike(f"%{q_filter}%"))
        if conf_filter:
            query = query.filter(QueryLog.confidence == conf_filter)
        if user_filter:
            query = query.filter(User.username == user_filter)

        pagination = query.paginate(page=page, per_page=30)
        return render_template(
            "admin/query_logs.html",
            pagination=pagination,
            q_filter=q_filter,
            conf_filter=conf_filter,
            user_filter=user_filter,
        )

    @admin_bp.route("/trace/<trace_id>")
    @login_required
    def trace_detail(trace_id):
        _require_admin()
        from models import QueryLog
        import json as _json

        ql = QueryLog.query.filter_by(trace_id=trace_id).first_or_404()
        user = User.query.get(ql.user_id) if ql.user_id else None
        repo = Repo.query.get(ql.repo_id)

        retrieval = {}
        if ql.retrieval_json:
            try:
                retrieval = _json.loads(ql.retrieval_json)
            except Exception:
                pass

        # list source files in the repo
        source_files = []
        if repo and user:
            repo_user = User.query.get(repo.user_id)
            repo_dir = os.path.join(Config.DATA_DIR, repo_user.username, repo.slug, "raw")
            if os.path.isdir(repo_dir):
                source_files = sorted(os.listdir(repo_dir))

        _, answer_html = render_markdown(ql.full_answer or "", "") if ql.full_answer else ("", "")

        return render_template(
            "admin/trace_detail.html",
            ql=ql, user=user, repo=repo,
            retrieval=retrieval,
            source_files=source_files,
            answer_html=answer_html,
        )

    @admin_bp.route("/repos")
    @login_required
    def all_repos():
        _require_admin()
        from models import QueryLog
        from sqlalchemy import func

        q_filter = request.args.get("q", "").strip()

        query = (
            db.session.query(
                Repo,
                User,
                func.count(QueryLog.id).label("query_count"),
            )
            .join(User, Repo.user_id == User.id)
            .outerjoin(QueryLog, QueryLog.repo_id == Repo.id)
            .group_by(Repo.id)
            .order_by(Repo.id.desc())
        )
        if q_filter:
            query = query.filter(
                Repo.name.ilike(f"%{q_filter}%") |
                Repo.slug.ilike(f"%{q_filter}%") |
                User.username.ilike(f"%{q_filter}%")
            )

        rows = query.all()
        return render_template("admin/repos.html", rows=rows, q_filter=q_filter)

    @admin_bp.route("/feedbacks")
    @login_required
    def feedbacks():
        _require_admin()
        from models import QueryFeedback, QueryLog
        page = request.args.get("page", 1, type=int)
        rating_filter = request.args.get("rating", "").strip()

        # MySQL：两表 trace_id 若 collation 不一致会报 Illegal mix of collations；SQLite 测试库无 COLLATE utf8mb4
        if db.engine.dialect.name == "mysql":
            trace_match = QueryFeedback.trace_id.collate(
                "utf8mb4_unicode_ci"
            ) == QueryLog.trace_id.collate("utf8mb4_unicode_ci")
        else:
            trace_match = QueryFeedback.trace_id == QueryLog.trace_id
        query = (
            db.session.query(QueryFeedback, User, Repo, QueryLog)
            .outerjoin(User, QueryFeedback.user_id == User.id)
            .join(Repo, QueryFeedback.repo_id == Repo.id)
            .outerjoin(QueryLog, trace_match)
            .order_by(QueryFeedback.created_at.desc())
        )
        if rating_filter:
            query = query.filter(QueryFeedback.rating == rating_filter)

        pagination = query.paginate(page=page, per_page=30)

        total_good = QueryFeedback.query.filter_by(rating="good").count()
        total_bad = QueryFeedback.query.filter_by(rating="bad").count()

        return render_template(
            "admin/feedbacks.html",
            pagination=pagination,
            rating_filter=rating_filter,
            total_good=total_good,
            total_bad=total_bad,
        )

    app.register_blueprint(admin_bp)


# ---------------------------------------------------------------------------
# SSE helper
# ---------------------------------------------------------------------------


def _sse(event: str, **data) -> str:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    application = create_app()
    with application.app_context():
        db.create_all()
    application.run(debug=True, host="0.0.0.0", port=5000)
